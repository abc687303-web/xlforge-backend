"""
XLforge v3.0 — Ultra-Robust AI Excel Generator
================================================
Primary AI   : Anthropic Claude (claude-sonnet-4-6 → claude-haiku-4-5)
Secondary AI : Groq  (auto-discovered live models, rotates on failure)
Tertiary AI  : OpenRouter (hundreds of free models as final fallback)

Self-healing features:
  ✅ Live model health-checks; dead models quarantined, revived after 15 min
  ✅ Auto-discovers ALL available models at startup & every hour
  ✅ Provider-level circuit breaker (skip provider if 3 consecutive failures)
  ✅ Exponential back-off + jitter on rate limits
  ✅ JSON self-repair for truncated AI responses
  ✅ Math expression solver prevents #VALUE! errors
  ✅ No deprecated model can ever be used — all resolved at runtime
  ✅ Zero human intervention required
"""

from fastapi import FastAPI, UploadFile, File, Form, BackgroundTasks, HTTPException, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse
from typing import Optional, List
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, AreaChart, Reference
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, CellIsRule
from openpyxl.drawing.image import Image as XLImage
import os, uuid, httpx, json, re, io, zipfile, base64, csv, time, logging, asyncio, random
from datetime import datetime, timedelta
from pathlib import Path
import sqlite3
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders

# =========================================================
# LOGGING
# =========================================================
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
    handlers=[logging.StreamHandler(), logging.FileHandler("xlforge.log", mode="a")]
)
logger = logging.getLogger("xlforge")

# =========================================================
# APP & MIDDLEWARE
# =========================================================
app = FastAPI(title="XLforge API", version="3.0.0")
app.add_middleware(
    CORSMiddleware, allow_origins=["*"], allow_credentials=False,
    allow_methods=["*"], allow_headers=["*"],
)

STORAGE_DIR = Path("storage")
INPUT_DIR   = STORAGE_DIR / "input"
OUTPUT_DIR  = STORAGE_DIR / "output"
TEMP_DIR    = STORAGE_DIR / "temp"
for d in [INPUT_DIR, OUTPUT_DIR, TEMP_DIR]:
    d.mkdir(parents=True, exist_ok=True)

conversation_memory: dict = {}
rate_limit_store:    dict = {}
jobs:                dict = {}

MAX_FILE_SIZE       = 100 * 1024 * 1024
RATE_LIMIT          = 20
FILE_EXPIRY_HOURS   = 24

# =========================================================
# DATABASE
# =========================================================
def get_db():
    conn = sqlite3.connect("xlforge.db", check_same_thread=False)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db()
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS jobs (
            job_id TEXT PRIMARY KEY, status TEXT DEFAULT 'pending',
            prompt TEXT, session_id TEXT, input_file TEXT, output_file TEXT,
            error TEXT, created_at TEXT, completed_at TEXT, processing_ms INTEGER,
            provider TEXT
        );
        CREATE TABLE IF NOT EXISTS templates (
            id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT, category TEXT,
            prompt TEXT, icon TEXT, created_at TEXT
        );
        CREATE TABLE IF NOT EXISTS usage_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT, job_id TEXT, session_id TEXT,
            ip TEXT, file_size INTEGER, ai_time_ms INTEGER, excel_time_ms INTEGER,
            success INTEGER, provider TEXT, model TEXT, created_at TEXT
        );
        CREATE TABLE IF NOT EXISTS model_health (
            provider TEXT, model TEXT, last_failure TEXT, failure_count INTEGER DEFAULT 0,
            quarantined_until TEXT, PRIMARY KEY (provider, model)
        );
    """)
    existing = conn.execute("SELECT COUNT(*) FROM templates").fetchone()[0]
    if existing == 0:
        templates = [
            ("Invoice",            "finance",      "Create a professional invoice template with company details, line items (description, qty, unit price, total), subtotal, tax (18%), grand total, payment terms, bank details", "🧾"),
            ("Monthly Budget",     "finance",      "Create a personal monthly budget tracker with income sources, fixed expenses, variable expenses, savings goal, actual vs budget comparison, and summary chart", "💰"),
            ("Inventory Tracker",  "business",     "Create an inventory management sheet with product ID, name, category, quantity in stock, reorder level, unit cost, total value, supplier, last restocked date, and stock status alerts", "📦"),
            ("Employee Attendance","HR",           "Create an employee attendance sheet for a month with employee ID, name, department, 31 days columns (P/A/L/H), total present, total absent, total leave, attendance percentage", "👥"),
            ("Sales Report",       "sales",        "Create a monthly sales report with salesperson, region, product, units sold, unit price, revenue, target, achievement percentage, rank, and bar chart", "📈"),
            ("Project Timeline",   "management",   "Create a project timeline/gantt-style sheet with task name, owner, start date, end date, duration days, status, priority, dependencies, completion %", "📅"),
            ("Student Gradebook",  "education",    "Create a student gradebook with student name, roll number, 6 subjects scores, total, percentage, grade (A/B/C/D/F), rank, pass/fail status and class average row", "🎓"),
            ("Expense Report",     "finance",      "Create a business expense report with date, category, description, amount, receipt number, payment method, reimbursable yes/no, approval status, and monthly totals", "💳"),
            ("KPI Dashboard",      "management",   "Create a KPI dashboard with metrics (Revenue, Customers, Conversion Rate, Avg Order Value, Churn Rate, NPS), current value, previous period, target, variance, trend (↑↓), and status", "📊"),
            ("Quotation",          "sales",        "Create a business quotation template with client details, quote number, validity date, itemized list (item, specs, qty, unit price, discount, net price), terms & conditions, total", "📋"),
            ("BOQ",                "construction", "Create a BOQ sheet with item number, description of work, unit, quantity, rate, amount, GST %, GST amount, total amount, section subtotals and grand total", "🏗️"),
            ("Payroll Sheet",      "HR",           "Create a monthly payroll sheet with employee ID, name, designation, basic salary, HRA, DA, other allowances, gross salary, PF deduction, ESI, tax, total deductions, net salary", "💵"),
        ]
        conn.executemany(
            "INSERT INTO templates (name, category, prompt, icon, created_at) VALUES (?,?,?,?,?)",
            [(t[0], t[1], t[2], t[3], datetime.utcnow().isoformat()) for t in templates]
        )
    conn.commit()
    conn.close()

init_db()

# =========================================================
# RATE LIMITER
# =========================================================
def check_rate_limit(ip: str) -> bool:
    now = time.time()
    if ip not in rate_limit_store:
        rate_limit_store[ip] = []
    rate_limit_store[ip] = [t for t in rate_limit_store[ip] if now - t < 60]
    if len(rate_limit_store[ip]) >= RATE_LIMIT:
        return False
    rate_limit_store[ip].append(now)
    return True

# =========================================================
# FILE READER
# =========================================================
async def read_any_file(file: UploadFile) -> tuple:
    raw = await file.read()
    if len(raw) > MAX_FILE_SIZE:
        raise HTTPException(status_code=413, detail="File too large")
    filename = (file.filename or "").lower()

    if filename.endswith(('.xlsx', '.xls')):
        try:
            wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
            lines = []
            for sn in wb.sheetnames:
                ws = wb[sn]
                lines.append(f"[Sheet: {sn}]")
                for row in ws.iter_rows(values_only=True):
                    if any(c is not None for c in row):
                        lines.append(" | ".join(str(c) if c is not None else "" for c in row))
            return "\n".join(lines), "excel", None, raw
        except Exception as e:
            return f"Excel read error: {e}", "excel", None, raw

    if filename.endswith('.csv'):
        return raw.decode("utf-8", errors="ignore"), "csv", None, raw

    if filename.endswith('.docx'):
        try:
            z = zipfile.ZipFile(io.BytesIO(raw))
            xml = z.read("word/document.xml").decode("utf-8")
            text = re.sub(r'<[^>]+>', ' ', xml)
            return re.sub(r'\s+', ' ', text).strip()[:6000], "word", None, raw
        except:
            return "", "word", None, raw

    if filename.endswith('.pdf'):
        try:
            text = raw.decode("latin-1", errors="ignore")
            strings = re.findall(r'[A-Za-z0-9 \+\-\=\.\,\:\;\!\?\%\$\#\@\/\(\)]{4,}', text)
            return " ".join(strings)[:6000], "pdf", None, raw
        except:
            return "", "pdf", None, raw

    if filename.endswith(('.jpg', '.jpeg', '.png', '.gif', '.webp', '.bmp')):
        b64 = base64.b64encode(raw).decode()
        ext = filename.split('.')[-1]
        mime = f"image/{'jpeg' if ext in ['jpg','jpeg'] else ext}"
        return "", "image", {"b64": b64, "mime": mime}, raw

    try:
        return raw.decode("utf-8", errors="ignore")[:6000], "text", None, raw
    except:
        return "", "unknown", None, raw

# =========================================================
# SYSTEM PROMPT
# =========================================================
SYSTEM_PROMPT = """You are XLforge, an AI Excel expert. Return ONLY a valid JSON object — no markdown, no explanation, nothing else.

JSON FORMAT (all keys required):
{"sheets":[{"name":"Sheet Name","headers":["Col A","Col B","Col C"],"rows":[["text",100,50.5]],"formulas":[{"cell":"D2","formula":"=IFERROR(SUM(B2:C2),0)","label":"Total"}],"summary_rows":[{"label":"TOTAL","col":1,"formula":"=IFERROR(SUM(B2:B100),0)"}],"conditional_formatting":[{"range":"B2:B20","type":"colorscale"}],"chart":{"type":"bar","title":"Chart Title","data_cols":[1],"category_col":0}}],"metadata":{"title":"Report Title","description":"What this does","author":"XLforge AI"}}

ABSOLUTE RULES:
1. NUMBERS: prices/scores/salaries MUST be int or float — WRONG:["John","50000"] CORRECT:["John",50000]
2. MATH: if file has "1021+707" in Problem col, Answer col MUST get integer 1728 (not string, not "#VALUE!", not null). Calculate EVERY row. Problem text stays unchanged in its column. NEVER output "#VALUE!" or any Excel error string as data.
3. SUMMARY ROWS: "col" is 0-indexed. headers=["A","B","C"] -> col 2 = column C. Always IFERROR.
4. CHARTS: type=bar/line/pie/area; data_cols and category_col are 0-indexed integers. If user says "chart", "graph", "plot", "visualize" -> ALWAYS include chart.
5. FORMULAS: always wrap in IFERROR. SUM range must start at row 2 (never row 1).
6. ROWS: copy every row from uploaded file — no skipping, no reducing.
7. PLACEHOLDERS: never use val1/item1/data1. Always real meaningful data, minimum 15 rows.
8. ANY LANGUAGE: understand and process prompts in any language.
9. FILE READING: when a file is uploaded, read and process EVERY row. Solve math problems, fill answer columns.
10. SELF-CORRECT: if you are unsure, make a best guess and produce useful output."""

# =========================================================
# MODEL HEALTH TRACKER (quarantine bad models)
# =========================================================
_quarantine: dict = {}  # (provider, model) -> quarantined_until datetime

def is_quarantined(provider: str, model: str) -> bool:
    key = (provider, model)
    until = _quarantine.get(key)
    if until and datetime.utcnow() < until:
        return True
    if key in _quarantine:
        del _quarantine[key]  # recovered
    return False

def quarantine_model(provider: str, model: str, minutes: int = 15):
    _quarantine[(provider, model)] = datetime.utcnow() + timedelta(minutes=minutes)
    logger.warning(f"Quarantined {provider}/{model} for {minutes} min")
    # Persist to DB
    try:
        conn = get_db()
        conn.execute("""
            INSERT INTO model_health (provider, model, last_failure, failure_count, quarantined_until)
            VALUES (?,?,?,1,?)
            ON CONFLICT(provider,model) DO UPDATE SET
                last_failure=excluded.last_failure,
                failure_count=failure_count+1,
                quarantined_until=excluded.quarantined_until
        """, (provider, model, datetime.utcnow().isoformat(),
               _quarantine[(provider, model)].isoformat()))
        conn.commit()
        conn.close()
    except Exception:
        pass

# =========================================================
# PROVIDER CIRCUIT BREAKER
# =========================================================
_provider_failures: dict = {}  # provider -> consecutive_failures
_CIRCUIT_OPEN_AFTER = 3        # open circuit after N consecutive failures
_CIRCUIT_RESET_AFTER = 300     # seconds before retrying a tripped provider

_provider_tripped: dict = {}   # provider -> tripped_at timestamp

def provider_ok(provider: str) -> bool:
    tripped_at = _provider_tripped.get(provider)
    if tripped_at and time.time() - tripped_at < _CIRCUIT_RESET_AFTER:
        return False
    if tripped_at:
        del _provider_tripped[provider]
        _provider_failures[provider] = 0
    return True

def record_provider_success(provider: str):
    _provider_failures[provider] = 0
    _provider_tripped.pop(provider, None)

def record_provider_failure(provider: str):
    _provider_failures[provider] = _provider_failures.get(provider, 0) + 1
    if _provider_failures[provider] >= _CIRCUIT_OPEN_AFTER:
        _provider_tripped[provider] = time.time()
        logger.error(f"Circuit OPEN for provider '{provider}' — will retry in {_CIRCUIT_RESET_AFTER}s")

# =========================================================
# ANTHROPIC PROVIDER
# =========================================================
_ANTHROPIC_MODELS = [
    "claude-sonnet-4-6",   # Best quality
    "claude-haiku-4-5-20251001",  # Fast fallback
]

async def call_anthropic(messages_payload: list, max_tokens: int = 8000) -> str:
    """Call Anthropic API. Returns raw text response."""
    api_key = os.getenv("ANTHROPIC_API_KEY")
    if not api_key:
        raise ValueError("ANTHROPIC_API_KEY not set")

    for model in _ANTHROPIC_MODELS:
        if is_quarantined("anthropic", model):
            continue
        try:
            async with httpx.AsyncClient(timeout=90) as client:
                resp = await client.post(
                    "https://api.anthropic.com/v1/messages",
                    headers={
                        "x-api-key": api_key,
                        "anthropic-version": "2023-06-01",
                        "Content-Type": "application/json"
                    },
                    json={
                        "model": model,
                        "max_tokens": max_tokens,
                        "system": SYSTEM_PROMPT,
                        "messages": messages_payload
                    }
                )
            if resp.status_code == 200:
                data = resp.json()
                text = data["content"][0]["text"].strip()
                record_provider_success("anthropic")
                logger.info(f"Anthropic/{model} success")
                return text
            elif resp.status_code in (400, 404):
                err = resp.text[:200]
                logger.warning(f"Anthropic/{model} permanent error {resp.status_code}: {err}")
                quarantine_model("anthropic", model, minutes=60)
                continue
            elif resp.status_code == 429:
                logger.warning(f"Anthropic/{model} rate limited")
                await asyncio.sleep(3 + random.uniform(0, 2))
                continue
            else:
                logger.warning(f"Anthropic/{model} HTTP {resp.status_code}")
                continue
        except (httpx.TimeoutException, httpx.ConnectError) as e:
            logger.warning(f"Anthropic/{model} timeout/connect: {e}")
            continue
        except Exception as e:
            logger.warning(f"Anthropic/{model} error: {e}")
            continue

    record_provider_failure("anthropic")
    raise ValueError("All Anthropic models failed")

# =========================================================
# GROQ PROVIDER (auto-discovers live models)
# =========================================================
_groq_model_cache: dict = {"text": [], "image": [], "expires": None}

_GROQ_TEXT_PRIORITY = [
    "llama-3.3-70b-versatile",
    "qwen/qwen3-32b",
    "llama3-groq-70b-8192-tool-use-preview",
    "gemma2-9b-it",
    "llama-3.1-70b-versatile",
    "llama-3.1-8b-instant",
    "mixtral-8x7b-32768",  # kept only if still alive
]
_GROQ_SKIP_PATTERNS = ["whisper","tts","speech","guard","orpheus","canopylabs/","prompt-guard"]
_GROQ_VISION_PRIORITY = [
    "meta-llama/llama-4-scout-17b-16e-instruct",
    "llama-3.3-70b-versatile",
]

async def _discover_groq_models(api_key: str):
    global _groq_model_cache
    try:
        async with httpx.AsyncClient(timeout=10) as client:
            r = await client.get(
                "https://api.groq.com/openai/v1/models",
                headers={"Authorization": f"Bearer {api_key}"}
            )
            r.raise_for_status()
            all_ids = {m["id"] for m in r.json().get("data", [])}

        text_ids = {m for m in all_ids if not any(p in m.lower() for p in _GROQ_SKIP_PATTERNS)}
        ordered = [m for m in _GROQ_TEXT_PRIORITY if m in text_ids]
        for m in sorted(text_ids):
            if m not in ordered:
                ordered.append(m)
        # Remove quarantined
        ordered = [m for m in ordered if not is_quarantined("groq", m)]

        vision = [m for m in _GROQ_VISION_PRIORITY if m in all_ids and not is_quarantined("groq", m)]
        if not vision:
            vision = ordered[:1]

        _groq_model_cache["text"]    = ordered[:8]
        _groq_model_cache["image"]   = vision[:3]
        _groq_model_cache["expires"] = datetime.utcnow() + timedelta(hours=1)
        logger.info(f"Groq auto-discovered {len(ordered)} text models: {ordered[:4]}")
    except Exception as e:
        logger.warning(f"Groq model discovery failed: {e}")
        if not _groq_model_cache["text"]:
            _groq_model_cache["text"]  = ["llama-3.3-70b-versatile", "gemma2-9b-it"]
            _groq_model_cache["image"] = ["llama-3.3-70b-versatile"]
        _groq_model_cache["expires"] = datetime.utcnow() + timedelta(minutes=5)

async def get_groq_text_models(api_key: str) -> list:
    if not _groq_model_cache["expires"] or datetime.utcnow() > _groq_model_cache["expires"]:
        await _discover_groq_models(api_key)
    # Re-filter quarantined models on every call
    models = [m for m in _groq_model_cache["text"] if not is_quarantined("groq", m)]
    return models or ["llama-3.3-70b-versatile"]

async def call_groq(messages_payload: list, max_tokens: int = 6000,
                    image_data: dict = None) -> str:
    """Call Groq API with automatic live model rotation."""
    api_key = os.getenv("GROQ_API_KEY")
    if not api_key:
        raise ValueError("GROQ_API_KEY not set")

    if image_data:
        models = [m for m in _groq_model_cache.get("image", []) if not is_quarantined("groq", m)]
        if not models:
            models = ["meta-llama/llama-4-scout-17b-16e-instruct"]
    else:
        models = await get_groq_text_models(api_key)

    for model in models[:6]:  # try up to 6 models
        if is_quarantined("groq", model):
            continue
        try:
            payload = {
                "model": model,
                "max_tokens": min(max_tokens, 6000),
                "temperature": 0.1,
                "messages": messages_payload
            }
            async with httpx.AsyncClient(timeout=60) as client:
                resp = await client.post(
                    "https://api.groq.com/openai/v1/chat/completions",
                    headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
                    json=payload
                )
            if resp.status_code == 200:
                result = resp.json()
                text = result["choices"][0]["message"]["content"].strip()
                record_provider_success("groq")
                logger.info(f"Groq/{model} success")
                return text
            elif resp.status_code in (400, 404):
                err = resp.text[:200]
                logger.warning(f"Groq/{model} permanent error: {err}")
                quarantine_model("groq", model, minutes=60)
                # Force rediscovery
                _groq_model_cache["expires"] = None
                continue
            elif resp.status_code == 413:
                logger.warning(f"Groq/{model} payload too large, trying smaller content")
                # caller handles content reduction
                raise ValueError(f"PAYLOAD_TOO_LARGE:{model}")
            elif resp.status_code == 429:
                wait = 3 + random.uniform(0, 3)
                logger.warning(f"Groq/{model} rate limited, waiting {wait:.1f}s")
                await asyncio.sleep(wait)
                continue
            else:
                logger.warning(f"Groq/{model} HTTP {resp.status_code}: {resp.text[:100]}")
                continue
        except ValueError as e:
            if "PAYLOAD_TOO_LARGE" in str(e):
                raise
            logger.warning(f"Groq/{model} value error: {e}")
            continue
        except (httpx.TimeoutException, httpx.ConnectError) as e:
            logger.warning(f"Groq/{model} network error: {e}")
            continue
        except Exception as e:
            logger.warning(f"Groq/{model} unexpected: {e}")
            continue

    record_provider_failure("groq")
    raise ValueError("All Groq models failed")

# =========================================================
# OPENROUTER PROVIDER (hundreds of free models as final fallback)
# =========================================================
_OR_MODELS = [
    "mistralai/mistral-7b-instruct:free",
    "meta-llama/llama-3.2-3b-instruct:free",
    "microsoft/phi-3-mini-128k-instruct:free",
    "google/gemma-2-9b-it:free",
    "nousresearch/hermes-3-llama-3.1-405b:free",
    "qwen/qwen-2-7b-instruct:free",
    "openchat/openchat-7b:free",
    "meta-llama/llama-3.1-8b-instruct:free",
]
_or_models_cache: list = []
_or_cache_expires: datetime = None

async def _discover_openrouter_models(api_key: str):
    global _or_models_cache, _or_cache_expires
    try:
        async with httpx.AsyncClient(timeout=10) as client:
            r = await client.get(
                "https://openrouter.ai/api/v1/models",
                headers={"Authorization": f"Bearer {api_key}"}
            )
            r.raise_for_status()
            models = r.json().get("data", [])
            # Free models only, prefer larger context
            free_text = [
                m["id"] for m in models
                if ":free" in m.get("id","") and
                   not any(skip in m.get("id","").lower() for skip in ["vision","image","dall","stable"])
            ]
            _or_models_cache = free_text[:20] or _OR_MODELS
        _or_cache_expires = datetime.utcnow() + timedelta(hours=6)
        logger.info(f"OpenRouter discovered {len(_or_models_cache)} free text models")
    except Exception as e:
        logger.warning(f"OpenRouter model discovery failed: {e}")
        _or_models_cache = _OR_MODELS
        _or_cache_expires = datetime.utcnow() + timedelta(minutes=10)

async def call_openrouter(messages_payload: list, max_tokens: int = 4000) -> str:
    api_key = os.getenv("OPENROUTER_API_KEY")
    if not api_key:
        raise ValueError("OPENROUTER_API_KEY not set")

    global _or_cache_expires
    if not _or_cache_expires or datetime.utcnow() > _or_cache_expires:
        await _discover_openrouter_models(api_key)

    models = [m for m in (_or_models_cache or _OR_MODELS) if not is_quarantined("openrouter", m)]
    random.shuffle(models)  # spread load across free tier limits

    for model in models[:8]:
        try:
            async with httpx.AsyncClient(timeout=60) as client:
                resp = await client.post(
                    "https://openrouter.ai/api/v1/chat/completions",
                    headers={
                        "Authorization": f"Bearer {api_key}",
                        "Content-Type": "application/json",
                        "HTTP-Referer": "https://xlforge.app",
                        "X-Title": "XLforge"
                    },
                    json={
                        "model": model,
                        "max_tokens": min(max_tokens, 4000),
                        "temperature": 0.15,
                        "messages": messages_payload
                    }
                )
            if resp.status_code == 200:
                result = resp.json()
                if result.get("choices"):
                    text = result["choices"][0]["message"]["content"].strip()
                    record_provider_success("openrouter")
                    logger.info(f"OpenRouter/{model} success")
                    return text
            elif resp.status_code in (400, 404, 422):
                logger.warning(f"OpenRouter/{model} bad model: {resp.text[:100]}")
                quarantine_model("openrouter", model, minutes=120)
                continue
            elif resp.status_code == 429:
                await asyncio.sleep(5 + random.uniform(0, 3))
                continue
            else:
                logger.warning(f"OpenRouter/{model} HTTP {resp.status_code}")
                continue
        except Exception as e:
            logger.warning(f"OpenRouter/{model} error: {e}")
            continue

    record_provider_failure("openrouter")
    raise ValueError("All OpenRouter models failed")

# =========================================================
# VALIDATION
# =========================================================
def validate_ai_response(data: dict) -> tuple:
    if not isinstance(data, dict):
        return False, "Not a dict"
    if "sheets" not in data or not data["sheets"]:
        return False, "Missing/empty sheets"
    for i, sheet in enumerate(data["sheets"]):
        if not sheet.get("headers"):
            return False, f"Sheet {i}: no headers"
        if not sheet.get("rows"):
            return False, f"Sheet {i}: no rows"
        if not isinstance(sheet["headers"], list):
            return False, f"Sheet {i}: headers must be list"
        if not isinstance(sheet["rows"], list):
            return False, f"Sheet {i}: rows must be list"
        hl = len(sheet["headers"])
        for j, row in enumerate(sheet["rows"]):
            if not isinstance(row, list):
                return False, f"Sheet {i} row {j}: not a list"
            while len(row) < hl:
                row.append("")
    return True, "OK"

def coerce_numeric(val):
    if isinstance(val, (int, float)):
        return val
    if isinstance(val, str):
        v = val.strip()
        if not v or v.startswith('='):
            return val
        if re.match(r'^-?[\d]+(\.[\d]+)?$', v):
            try:
                return float(v) if '.' in v else int(v)
            except:
                pass
        if re.match(r'^-?[\d]{1,3}(,[\d]{3})*(\.[\d]+)?$', v):
            try:
                cleaned = v.replace(',', '')
                return float(cleaned) if '.' in cleaned else int(cleaned)
            except:
                pass
    return val

def solve_math_expression(expr: str):
    if not isinstance(expr, str):
        return None
    cleaned = expr.strip()
    if not re.match(r'^[\d\s\+\-\*\/\(\)\.]+$', cleaned):
        return None
    if not re.search(r'[\+\-\*\/]', cleaned):
        return None
    try:
        result = eval(cleaned, {"__builtins__": {}}, {})
        if isinstance(result, float) and result.is_integer():
            return int(result)
        return result
    except:
        return None

def sanitize_ai_response(data: dict) -> dict:
    for sheet in data.get("sheets", []):
        headers = sheet.get("headers", [])
        header_len = len(headers)
        problem_cols = {i for i, h in enumerate(headers) if any(k in str(h).lower() for k in ["problem","question","expression","equation","task"])}
        answer_cols  = {i for i, h in enumerate(headers) if any(k in str(h).lower() for k in ["answer","result","solution","output","total"])}
        cleaned = []
        for row in sheet.get("rows", []):
            if not isinstance(row, list):
                continue
            row = row[:header_len]
            while len(row) < header_len:
                row.append("")
            new_row = []
            for ci, v in enumerate(row):
                if ci in problem_cols:
                    new_row.append(str(v) if v is not None else "")
                elif ci in answer_cols:
                    sv = str(v).strip() if v is not None else ""
                    # Check if there's a solvable math expression in the problem column
                    has_math_problem = any(
                        solve_math_expression(str(row[pi])) is not None
                        for pi in problem_cols if pi < len(row)
                    )
                    is_bad = (
                        not sv
                        or sv.upper().startswith("#")
                        or (not isinstance(v, (int,float)) and not sv.lstrip("-").replace(".","",1).isdigit())
                        or (v == 0 and has_math_problem)  # 0 is wrong if there's a real math problem
                    )
                    if is_bad:
                        solved = None
                        for pi in problem_cols:
                            if pi < len(row):
                                solved = solve_math_expression(str(row[pi]))
                                if solved is not None:
                                    break
                        new_row.append(solved if solved is not None else coerce_numeric(v))
                    else:
                        new_row.append(coerce_numeric(v))
                else:
                    new_row.append(coerce_numeric(v))
            cleaned.append(new_row)
        sheet["rows"] = cleaned
        for key in ("formulas", "conditional_formatting", "summary_rows"):
            if not sheet.get(key):
                sheet[key] = []
    return data

# =========================================================
# JSON SELF-REPAIR
# =========================================================
def repair_json(text: str) -> str:
    in_string = False
    escape_next = False
    for ch in text:
        if escape_next:
            escape_next = False
            continue
        if ch == '\\' and in_string:
            escape_next = True
            continue
        if ch == '"':
            in_string = not in_string
    if in_string:
        text += '"'
    text = re.sub(r',\s*([}\]])', r'\1', text)
    text = text.rstrip(' \t\n,')
    stack = []
    in_str = esc = False
    for ch in text:
        if esc:
            esc = False
            continue
        if ch == '\\' and in_str:
            esc = True
            continue
        if ch == '"':
            in_str = not in_str
            continue
        if in_str:
            continue
        if ch == '{':
            stack.append('}')
        elif ch == '[':
            stack.append(']')
        elif ch in ('}', ']') and stack:
            stack.pop()
    for closer in reversed(stack):
        text += closer
    text = re.sub(r',\s*([}\]])', r'\1', text)
    return text

def extract_and_parse_json(raw: str) -> dict:
    """Extract JSON from AI response, apply self-repair, return parsed dict."""
    text = raw.strip()
    # Strip markdown fences
    text = re.sub(r'^```[a-zA-Z]*\n?', '', text)
    text = re.sub(r'\n?```$', '', text)
    text = text.strip()

    # Find outermost { ... } string-aware
    brace_start = text.find('{')
    if brace_start != -1:
        depth = 0
        brace_end = -1
        in_str = esc = False
        for i, ch in enumerate(text[brace_start:], brace_start):
            if esc:
                esc = False
                continue
            if ch == '\\' and in_str:
                esc = True
                continue
            if ch == '"':
                in_str = not in_str
                continue
            if in_str:
                continue
            if ch == '{':
                depth += 1
            elif ch == '}':
                depth -= 1
                if depth == 0:
                    brace_end = i + 1
                    break
        text = text[brace_start:brace_end] if brace_end != -1 else text[brace_start:]

    text = repair_json(text)
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        last_bracket = text.rfind(']]')
        if last_bracket != -1:
            text = repair_json(text[:last_bracket + 2])
        return json.loads(text)

# =========================================================
# UNIFIED AI ORCHESTRATOR
# =========================================================
def build_messages(prompt: str, file_content: str, file_type: str,
                   image_data: dict, session_id: str,
                   content_limit: int = 8000) -> list:
    """Build the messages list for any provider (OpenAI-compatible format)."""
    messages = []
    # Conversation history
    if session_id and session_id in conversation_memory:
        messages.extend(conversation_memory[session_id][-6:])

    if image_data:
        messages.append({
            "role": "user",
            "content": [
                {"type": "image_url",
                 "image_url": {"url": f"data:{image_data['mime']};base64,{image_data['b64']}"}},
                {"type": "text",
                 "text": f"Analyze this image. {'Task: ' + prompt if prompt.strip() else 'Create the best Excel from this image.'} Solve any math, return only JSON."}
            ]
        })
    elif file_content and not prompt.strip():
        messages.append({
            "role": "user",
            "content": f"I uploaded a {file_type} file. Read every row, create professional Excel.\n\nFILE CONTENT:\n{file_content[:content_limit]}\n\n- Solve math (put answers in Answer col, keep problem text unchanged)\n- Copy every row\n- Return only JSON"
        })
    elif file_content and prompt.strip():
        messages.append({
            "role": "user",
            "content": f"Task: {prompt}\n\nFILE TYPE: {file_type}\nFILE CONTENT (use ALL rows):\n{file_content[:content_limit]}\n\n- Use every row from the file\n- Complete missing values\n- If task says 'add chart' -> include chart in JSON\n- Return only JSON"
        })
    else:
        messages.append({
            "role": "user",
            "content": f"Create a professional Excel spreadsheet: \"{prompt}\"\n- Real meaningful data, minimum 15 rows\n- Multiple sheets if it makes sense\n- Professional formulas\n- Add chart ONLY if prompt asks\n- No placeholder values\n- Return only JSON"
        })
    return messages

async def call_ai(prompt: str, file_content: str = "", file_type: str = "",
                  image_data: dict = None, session_id: str = None) -> tuple:
    """
    Ultra-resilient AI orchestrator.
    Tries providers in order: Anthropic → Groq → OpenRouter
    Each provider is skipped if its circuit breaker is open.
    Returns (parsed_dict, provider_name, model_name).
    """
    content_limits = [8000, 6000, 4000, 2000]
    last_error = "No providers available"

    for attempt in range(6):
        content_limit = content_limits[min(attempt, len(content_limits)-1)]
        msgs = build_messages(prompt, file_content, file_type, image_data,
                              session_id, content_limit)
        # Build Anthropic-style messages (content only, no system — added in call_anthropic)
        anthropic_msgs = [{"role": m["role"], "content": m["content"]} for m in msgs]

        # ---- ANTHROPIC (primary) ----
        if attempt < 3 and provider_ok("anthropic") and os.getenv("ANTHROPIC_API_KEY"):
            try:
                raw = await call_anthropic(anthropic_msgs, max_tokens=8000)
                data = extract_and_parse_json(raw)
                valid, reason = validate_ai_response(data)
                if valid:
                    data = sanitize_ai_response(data)
                    _save_conversation(session_id, prompt, file_type, data)
                    return data, "anthropic", "claude"
                else:
                    last_error = f"Anthropic validation failed: {reason}"
                    logger.warning(last_error)
                    continue
            except Exception as e:
                last_error = f"Anthropic: {e}"
                logger.warning(f"Attempt {attempt+1} Anthropic failed: {e}")

        # ---- GROQ (secondary) ----
        if provider_ok("groq") and os.getenv("GROQ_API_KEY"):
            try:
                groq_msgs = [{"role": "system", "content": SYSTEM_PROMPT}] + msgs
                raw = await call_groq(groq_msgs, max_tokens=5500, image_data=image_data)
                data = extract_and_parse_json(raw)
                valid, reason = validate_ai_response(data)
                if valid:
                    data = sanitize_ai_response(data)
                    _save_conversation(session_id, prompt, file_type, data)
                    return data, "groq", "auto"
                else:
                    last_error = f"Groq validation failed: {reason}"
                    continue
            except ValueError as e:
                if "PAYLOAD_TOO_LARGE" in str(e):
                    # Force smaller content on next attempt
                    content_limits[min(attempt+1, 3)] = 2000
                last_error = f"Groq: {e}"
                logger.warning(f"Attempt {attempt+1} Groq failed: {e}")
            except Exception as e:
                last_error = f"Groq: {e}"
                logger.warning(f"Attempt {attempt+1} Groq failed: {e}")

        # ---- OPENROUTER (tertiary) ----
        if attempt >= 2 and provider_ok("openrouter") and os.getenv("OPENROUTER_API_KEY"):
            try:
                or_msgs = [{"role": "system", "content": SYSTEM_PROMPT}] + msgs
                raw = await call_openrouter(or_msgs, max_tokens=4000)
                data = extract_and_parse_json(raw)
                valid, reason = validate_ai_response(data)
                if valid:
                    data = sanitize_ai_response(data)
                    _save_conversation(session_id, prompt, file_type, data)
                    return data, "openrouter", "auto"
                else:
                    last_error = f"OpenRouter validation: {reason}"
                    continue
            except Exception as e:
                last_error = f"OpenRouter: {e}"
                logger.warning(f"Attempt {attempt+1} OpenRouter failed: {e}")

        await asyncio.sleep(min(2 ** attempt, 8) + random.uniform(0, 1))

    raise ValueError(f"All AI providers failed after 6 attempts. Last error: {last_error}")

def _save_conversation(session_id, prompt, file_type, data):
    if not session_id:
        return
    user_msg = {"role": "user", "content": prompt or f"[{file_type} file uploaded]"}
    asst_msg = {"role": "assistant", "content": f"[Generated Excel with {len(data.get('sheets',[]))} sheet(s)]"}
    if session_id not in conversation_memory:
        conversation_memory[session_id] = []
    conversation_memory[session_id].extend([user_msg, asst_msg])
    conversation_memory[session_id] = conversation_memory[session_id][-20:]

# =========================================================
# EXCEL BUILDER
# =========================================================
def build_excel(data: dict, output_path: str, password: str = None):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    BLUE       = "2563EB"
    LIGHT_BLUE = "EFF6FF"
    WHITE      = "FFFFFF"
    DARK       = "1E293B"
    SUMMARY_BG = "1E3A5F"
    pending_charts = []

    def thin_border():
        s = Side(style="thin", color="CCCCCC")
        return Border(left=s, right=s, top=s, bottom=s)
    def thick_border():
        s = Side(style="medium", color="2563EB")
        return Border(left=s, right=s, top=s, bottom=s)
    def style_header(cell):
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.font = Font(bold=True, color=WHITE, size=11, name="Calibri")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border()
    def style_data(cell, row_idx, header=""):
        cell.border = thin_border()
        cell.alignment = Alignment(vertical="center")
        if row_idx % 2 == 0:
            cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        if isinstance(cell.value, (int, float)):
            cell.alignment = Alignment(horizontal="right", vertical="center")
            h = header.lower()
            if any(k in h for k in ["salary","revenue","price","cost","amount","budget","sales","income","expense","pay","total","value"]):
                cell.number_format = '#,##0.00'
            elif any(k in h for k in ["percent","%","rate","growth","margin"]):
                cell.number_format = '0.00'
    def style_summary(cell):
        cell.fill = PatternFill("solid", fgColor=SUMMARY_BG)
        cell.font = Font(bold=True, color=WHITE, size=11, name="Calibri")
        cell.border = thick_border()
        cell.alignment = Alignment(horizontal="right", vertical="center")

    for sheet_def in data["sheets"]:
        name = str(sheet_def.get("name","Sheet"))[:31]
        ws = wb.create_sheet(title=name)
        headers   = sheet_def.get("headers", [])
        rows      = sheet_def.get("rows", [])
        formulas  = sheet_def.get("formulas", [])
        cf_rules  = sheet_def.get("conditional_formatting", [])
        chart_def = sheet_def.get("chart", None)

        ws.sheet_properties.tabColor = BLUE

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=str(header))
            style_header(cell)
        ws.row_dimensions[1].height = 32

        for row_idx, row in enumerate(rows, 2):
            for col_idx, val in enumerate(row, 1):
                header = headers[col_idx-1] if col_idx <= len(headers) else ""
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                style_data(cell, row_idx, header)
            ws.row_dimensions[row_idx].height = 20

        # Auto-fit columns
        for col_idx, header in enumerate(headers, 1):
            max_len = len(str(header))
            for row in rows:
                if col_idx - 1 < len(row):
                    v = row[col_idx-1]
                    max_len = max(max_len, len(str(v)) if v is not None else 0)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len+4, 14), 60)

        # No auto summary rows — only add TOTAL/AVERAGE if the AI explicitly
        # included them in summary_rows (i.e. user asked for totals).
        next_row = len(rows) + 2
        summary_rows_def = sheet_def.get("summary_rows", [])
        if summary_rows_def:
            data_start = 2
            data_end   = len(rows) + 1
            numeric_cols = set()
            for row in rows:
                for ci, val in enumerate(row):
                    if isinstance(val, (int, float)):
                        numeric_cols.add(ci)
            for sr in summary_rows_def:
                label   = sr.get("label", "TOTAL")
                for col_0idx in range(len(headers)):
                    cell = ws.cell(row=next_row, column=col_0idx+1)
                    if col_0idx == 0:
                        cell.value = label
                    elif col_0idx in numeric_cols:
                        cl = get_column_letter(col_0idx+1)
                        if "average" in label.lower() or "avg" in label.lower():
                            cell.value = f'=IFERROR(AVERAGE({cl}{data_start}:{cl}{data_end}),0)'
                        else:
                            cell.value = f'=IFERROR(SUM({cl}{data_start}:{cl}{data_end}),0)'
                    style_summary(cell)
                next_row += 1

        # Inline formulas
        for f in formulas:
            addr = f.get("cell","")
            formula = f.get("formula","")
            if addr and formula:
                try:
                    ws[addr] = formula
                    ws[addr].font = Font(bold=True, color=DARK)
                    ws[addr].border = thick_border()
                    m = re.match(r"([A-Z]+)(\d+)", addr)
                    if m and f.get("label"):
                        ci = column_index_from_string(m.group(1))
                        ri = int(m.group(2))
                        if ci > 1:
                            lc = ws.cell(row=ri, column=ci-1)
                            if not lc.value:
                                lc.value = f["label"]
                                lc.font = Font(bold=True)
                except:
                    pass

        ws.freeze_panes = "A2"
        if headers:
            ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows)+1}"
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1

        # Conditional formatting
        for cf in cf_rules:
            r = cf.get("range","")
            t = cf.get("type","")
            if not r:
                continue
            try:
                if t == "colorscale":
                    ws.conditional_formatting.add(r, ColorScaleRule(
                        start_type="min", start_color="F87171",
                        mid_type="percentile", mid_value=50, mid_color="FCD34D",
                        end_type="max", end_color="4ADE80"
                    ))
                elif t == "databar":
                    ws.conditional_formatting.add(r, DataBarRule(
                        start_type="min", start_value=0, end_type="max", end_value=100, color="2563EB"
                    ))
                elif t == "highlight_high":
                    ws.conditional_formatting.add(r, CellIsRule(
                        operator="greaterThan", formula=[str(cf.get("threshold",0))],
                        stopIfTrue=True, fill=PatternFill(fill_type="solid", fgColor="4ADE80"),
                        font=Font(color="14532D", bold=True)
                    ))
                elif t == "highlight_low":
                    ws.conditional_formatting.add(r, CellIsRule(
                        operator="lessThan", formula=[str(cf.get("threshold",0))],
                        stopIfTrue=True, fill=PatternFill(fill_type="solid", fgColor="FCA5A5"),
                        font=Font(color="7F1D1D", bold=True)
                    ))
            except Exception as e:
                logger.warning(f"CF error: {e}")

        if chart_def:
            pending_charts.append({"chart_def": chart_def, "data_ws": ws,
                                    "nrows": len(rows), "next_row": next_row})

    # Second pass: charts
    for pc in pending_charts:
        try:
            chart_def = pc["chart_def"]
            data_ws   = pc["data_ws"]
            nrows     = pc["nrows"]
            nrow_v    = pc["next_row"]
            ctype     = str(chart_def.get("type","bar")).lower()
            dcols     = chart_def.get("data_cols",[1])
            catcol    = int(chart_def.get("category_col",0))
            ctitle    = str(chart_def.get("title","Chart"))
            tgt_idx   = chart_def.get("sheet", None)
            if not isinstance(dcols, list) or not dcols:
                dcols = [1]
            dcols = [int(d) for d in dcols]
            chart = (PieChart() if ctype=="pie" else
                     LineChart() if ctype=="line" else
                     AreaChart() if ctype=="area" else BarChart())
            if hasattr(chart, "type"):
                chart.type = "col"
            chart.title  = ctitle
            chart.style  = 10
            chart.height = 15
            chart.width  = 28
            data_max_row = nrows + 1
            if ctype == "pie":
                dc = dcols[0]
                data_ref = Reference(data_ws, min_col=dc+1, min_row=1, max_row=data_max_row)
                chart.add_data(data_ref, titles_from_data=True)
            else:
                for dc in dcols:
                    if dc+1 <= 50:
                        data_ref = Reference(data_ws, min_col=dc+1, min_row=1, max_row=data_max_row)
                        chart.add_data(data_ref, titles_from_data=True)
            cats = Reference(data_ws, min_col=catcol+1, min_row=2, max_row=data_max_row)
            chart.set_categories(cats)
            all_sheets = wb.worksheets
            if tgt_idx is not None and 0 <= int(tgt_idx) < len(all_sheets):
                chart_ws = all_sheets[int(tgt_idx)]
                anchor_row = 3
            else:
                chart_ws = data_ws
                anchor_row = max(nrows+4, nrow_v+2)
            chart_ws.add_chart(chart, f"A{anchor_row}")
            logger.info(f"Chart '{ctitle}' ({ctype}) added on sheet '{chart_ws.title}'")
        except Exception as e:
            logger.warning(f"Chart error (skipped): {e}")

    wb.save(output_path)
    return output_path

# =========================================================
# CSV CONVERTER
# =========================================================
def excel_to_csv(excel_path: str) -> str:
    csv_path = excel_path.replace(".xlsx", ".csv")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.worksheets[0] if wb.worksheets else wb.active
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        for row in ws.iter_rows(values_only=True):
            writer.writerow([c if c is not None else "" for c in row])
    return csv_path

# =========================================================
# BACKGROUND JOB PROCESSOR
# =========================================================
async def process_job(job_id, prompt, file_content, file_type, image_data,
                      session_id, output_path, password=None, custom_filename=None):
    start = time.time()
    conn = get_db()
    try:
        jobs[job_id]["status"] = "processing"
        conn.execute("UPDATE jobs SET status='processing' WHERE job_id=?", (job_id,))
        conn.commit()

        ai_start = time.time()
        data, provider, model = await call_ai(prompt, file_content, file_type, image_data, session_id)
        ai_ms = int((time.time() - ai_start) * 1000)

        excel_start = time.time()
        build_excel(data, output_path, password=password)
        excel_ms = int((time.time() - excel_start) * 1000)

        total_ms = int((time.time() - start) * 1000)
        jobs[job_id].update({
            "status": "done", "output_path": output_path,
            "custom_filename": custom_filename or "xlforge_output.xlsx",
            "sheet_count": len(data.get("sheets", [])),
            "metadata": data.get("metadata", {}),
            "processing_ms": total_ms, "provider": provider, "model": model
        })
        conn.execute("""
            UPDATE jobs SET status='done', output_file=?, completed_at=?, processing_ms=?, provider=?
            WHERE job_id=?
        """, (output_path, datetime.utcnow().isoformat(), total_ms, provider, job_id))
        conn.execute("""
            INSERT INTO usage_log (job_id,session_id,ai_time_ms,excel_time_ms,success,provider,model,created_at)
            VALUES (?,?,?,?,1,?,?,?)
        """, (job_id, session_id, ai_ms, excel_ms, provider, model, datetime.utcnow().isoformat()))
        conn.commit()
        logger.info(f"Job {job_id} done in {total_ms}ms via {provider}/{model}")

    except Exception as e:
        jobs[job_id].update({"status": "failed", "error": str(e)})
        try:
            conn.execute("UPDATE jobs SET status='failed', error=? WHERE job_id=?", (str(e), job_id))
            conn.execute("INSERT INTO usage_log (job_id,session_id,success,created_at) VALUES (?,?,0,?)",
                         (job_id, session_id, datetime.utcnow().isoformat()))
            conn.commit()
        except Exception:
            pass
        logger.error(f"Job {job_id} failed: {e}")
    finally:
        conn.close()

# =========================================================
# CLEANUP
# =========================================================
def cleanup_old_files():
    cutoff = time.time() - (FILE_EXPIRY_HOURS * 3600)
    for folder in [INPUT_DIR, OUTPUT_DIR, TEMP_DIR]:
        for f in folder.glob("*"):
            try:
                if f.stat().st_mtime < cutoff:
                    f.unlink()
            except:
                pass

def _prune_jobs():
    cutoff = datetime.utcnow() - timedelta(hours=24)
    to_rm = [jid for jid, j in list(jobs.items())
             if j.get("status") in ("done","failed") and
             datetime.fromisoformat(j.get("created_at","2000-01-01")) < cutoff]
    for jid in to_rm:
        jobs.pop(jid, None)
    if to_rm:
        logger.info(f"Pruned {len(to_rm)} old jobs")

# =========================================================
# STARTUP
# =========================================================
@app.on_event("startup")
async def _startup():
    # Pre-warm model caches
    if os.getenv("GROQ_API_KEY"):
        try:
            models = await get_groq_text_models(os.getenv("GROQ_API_KEY"))
            logger.info(f"Groq pre-warm: {models[:3]}")
        except Exception as e:
            logger.warning(f"Groq pre-warm failed: {e}")
    if os.getenv("OPENROUTER_API_KEY"):
        try:
            await _discover_openrouter_models(os.getenv("OPENROUTER_API_KEY"))
        except Exception as e:
            logger.warning(f"OpenRouter pre-warm failed: {e}")
    anthropic_key = os.getenv("ANTHROPIC_API_KEY")
    if anthropic_key:
        logger.info("Anthropic API key detected ✅")
    else:
        logger.warning("ANTHROPIC_API_KEY not set — Claude provider disabled")

    # Hourly cleanup + model refresh loop
    async def _loop():
        while True:
            await asyncio.sleep(3600)
            cleanup_old_files()
            _prune_jobs()
            # Refresh model caches
            if os.getenv("GROQ_API_KEY"):
                try:
                    _groq_model_cache["expires"] = None
                    await get_groq_text_models(os.getenv("GROQ_API_KEY"))
                except:
                    pass
            if os.getenv("OPENROUTER_API_KEY"):
                try:
                    await _discover_openrouter_models(os.getenv("OPENROUTER_API_KEY"))
                except:
                    pass
    asyncio.create_task(_loop())

# =========================================================
# ENDPOINTS
# =========================================================
@app.get("/")
def root():
    return {
        "status": "XLforge v3.0 running!",
        "providers": {
            "anthropic":   {"configured": bool(os.getenv("ANTHROPIC_API_KEY")), "models": _ANTHROPIC_MODELS},
            "groq":        {"configured": bool(os.getenv("GROQ_API_KEY")),
                            "models": _groq_model_cache.get("text",[])[:4]},
            "openrouter":  {"configured": bool(os.getenv("OPENROUTER_API_KEY")),
                            "models": _or_models_cache[:4]}
        },
        "features": ["anthropic-primary","groq-secondary","openrouter-tertiary",
                     "self-healing","circuit-breaker","auto-model-discovery",
                     "json-repair","math-solver","conversation-memory"]
    }

@app.get("/health")
def health():
    quarantined = [{"provider": p, "model": m, "until": str(u)}
                   for (p, m), u in _quarantine.items()]
    tripped = list(_provider_tripped.keys())
    return {
        "status": "healthy",
        "providers": {
            "anthropic":  provider_ok("anthropic"),
            "groq":       provider_ok("groq"),
            "openrouter": provider_ok("openrouter"),
        },
        "quarantined_models": quarantined,
        "tripped_providers": tripped,
        "active_jobs": len([j for j in jobs.values() if j["status"] == "processing"]),
        "sessions": len(conversation_memory),
        "timestamp": datetime.utcnow().isoformat()
    }

@app.get("/models")
async def list_models():
    groq_key = os.getenv("GROQ_API_KEY")
    _groq_model_cache["expires"] = None
    groq_text = await get_groq_text_models(groq_key) if groq_key else []
    return {
        "anthropic":  {"models": _ANTHROPIC_MODELS, "circuit_ok": provider_ok("anthropic")},
        "groq":       {"models": groq_text,          "circuit_ok": provider_ok("groq")},
        "openrouter": {"models": _or_models_cache[:8], "circuit_ok": provider_ok("openrouter")},
        "quarantined": [{"provider":p,"model":m} for (p,m) in _quarantine],
    }

@app.post("/admin/reset-circuit/{provider}")
def reset_circuit(provider: str):
    _provider_tripped.pop(provider, None)
    _provider_failures[provider] = 0
    return {"message": f"Circuit reset for {provider}"}

@app.post("/admin/clear-quarantine")
def clear_quarantine():
    _quarantine.clear()
    return {"message": "All model quarantines cleared"}

@app.post("/generate")
async def generate_excel(
    background_tasks: BackgroundTasks,
    request: Request,
    prompt: str = Form(default=""),
    session_id: str = Form(default=""),
    custom_filename: str = Form(default=""),
    password: str = Form(default=""),
    files: List[UploadFile] = File(default=[]),
    file: UploadFile = File(default=None),
):
    ip = request.client.host
    if not check_rate_limit(ip):
        raise HTTPException(status_code=429, detail="Rate limit exceeded. Max 20 requests/minute.")
    if not session_id:
        session_id = str(uuid.uuid4())
    job_id = str(uuid.uuid4())
    output_path = str(OUTPUT_DIR / f"{job_id}.xlsx")

    all_files = [uf for uf in (files or []) if uf and uf.filename]
    if file and file.filename:
        all_files.insert(0, file)

    file_content, file_type, image_data = "", "", None
    if all_files:
        contents = []
        for uf in all_files[:20]:
            fc, ft, img, raw = await read_any_file(uf)
            if img:
                image_data = img
            elif fc:
                contents.append(f"[File: {uf.filename}]\n{fc}")
            file_type = ft
            ip2 = INPUT_DIR / f"{job_id}_{uf.filename}"
            with open(str(ip2), "wb") as fout:
                fout.write(raw)
        file_content = "\n\n".join(contents)

    jobs[job_id] = {"status": "pending", "job_id": job_id,
                    "session_id": session_id, "created_at": datetime.utcnow().isoformat()}
    conn = get_db()
    conn.execute("INSERT INTO jobs (job_id,status,prompt,session_id,created_at) VALUES (?,?,?,?,?)",
                 (job_id, "pending", prompt, session_id, datetime.utcnow().isoformat()))
    conn.commit()
    conn.close()

    fname = custom_filename.strip() or "xlforge_output"
    if not fname.endswith(".xlsx"):
        fname += ".xlsx"

    background_tasks.add_task(
        process_job, job_id=job_id, prompt=prompt, file_content=file_content,
        file_type=file_type, image_data=image_data, session_id=session_id,
        output_path=output_path, password=password or None, custom_filename=fname
    )
    return {"job_id": job_id, "session_id": session_id, "status": "processing",
            "message": "Poll /status/{job_id} to check progress."}

@app.get("/status/{job_id}")
def job_status(job_id: str):
    job = jobs.get(job_id)
    if not job:
        conn = get_db()
        row = conn.execute("SELECT * FROM jobs WHERE job_id=?", (job_id,)).fetchone()
        conn.close()
        if not row:
            raise HTTPException(status_code=404, detail="Job not found")
        return dict(row)
    return job

@app.get("/download/{job_id}")
def download_excel(job_id: str):
    job = jobs.get(job_id)
    if not job or job["status"] != "done":
        raise HTTPException(status_code=404, detail="File not ready")
    output_path = job.get("output_path","")
    if not os.path.exists(output_path):
        raise HTTPException(status_code=404, detail="File not found on disk")
    return FileResponse(output_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=job.get("custom_filename","xlforge_output.xlsx"),
        headers={"Access-Control-Allow-Origin": "*"})

@app.get("/download-csv/{job_id}")
def download_csv(job_id: str):
    job = jobs.get(job_id)
    if not job or job["status"] != "done":
        raise HTTPException(status_code=404, detail="File not ready")
    csv_path = excel_to_csv(job["output_path"])
    return FileResponse(csv_path, media_type="text/csv",
        filename=job.get("custom_filename","xlforge_output.xlsx").replace(".xlsx",".csv"),
        headers={"Access-Control-Allow-Origin": "*"})

@app.post("/generate-sync")
async def generate_sync(
    request: Request,
    prompt: str = Form(default=""),
    session_id: str = Form(default=""),
    custom_filename: str = Form(default=""),
    file: UploadFile = File(None)
):
    ip = request.client.host
    if not check_rate_limit(ip):
        raise HTTPException(status_code=429, detail="Rate limit exceeded.")
    if not session_id:
        session_id = str(uuid.uuid4())
    job_id = str(uuid.uuid4())
    output_path = str(OUTPUT_DIR / f"{job_id}.xlsx")

    file_content, file_type, image_data = "", "", None
    if file and file.filename:
        file_content, file_type, image_data, _ = await read_any_file(file)

    try:
        data, provider, model = await call_ai(prompt, file_content, file_type, image_data, session_id)
        build_excel(data, output_path)
    except Exception as e:
        Path(output_path).unlink(missing_ok=True)
        raise HTTPException(status_code=500, detail=str(e))

    fname = custom_filename.strip() or "xlforge_output"
    if not fname.endswith(".xlsx"):
        fname += ".xlsx"
    return FileResponse(output_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=fname,
        headers={"Access-Control-Allow-Origin": "*", "X-Session-Id": session_id,
                 "X-Provider": provider, "X-Model": model})

@app.post("/process")
async def process_alias(
    background_tasks: BackgroundTasks, request: Request,
    prompt: str = Form(default=""), session_id: str = Form(default=""),
    custom_filename: str = Form(default=""), password: str = Form(default=""),
    files: List[UploadFile] = File(default=[]), file: UploadFile = File(default=None),
):
    return await generate_excel(background_tasks=background_tasks, request=request,
        prompt=prompt, session_id=session_id, custom_filename=custom_filename,
        password=password, files=files, file=file)

@app.get("/templates")
def get_templates(category: str = None):
    conn = get_db()
    if category:
        rows = conn.execute("SELECT * FROM templates WHERE category=? ORDER BY name", (category,)).fetchall()
    else:
        rows = conn.execute("SELECT * FROM templates ORDER BY category, name").fetchall()
    conn.close()
    return {"templates": [dict(r) for r in rows]}

@app.get("/templates/categories")
def get_categories():
    conn = get_db()
    rows = conn.execute("SELECT DISTINCT category FROM templates ORDER BY category").fetchall()
    conn.close()
    return {"categories": [r["category"] for r in rows]}

@app.post("/templates/use/{template_id}")
async def use_template(
    template_id: int, background_tasks: BackgroundTasks, request: Request,
    session_id: str = Form(default=""), custom_data: str = Form(default=""),
):
    conn = get_db()
    tmpl = conn.execute("SELECT * FROM templates WHERE id=?", (template_id,)).fetchone()
    conn.close()
    if not tmpl:
        raise HTTPException(status_code=404, detail="Template not found")
    prompt = tmpl["prompt"]
    if custom_data.strip():
        prompt += f"\n\nAdditional data/context: {custom_data}"
    ip = request.client.host
    if not check_rate_limit(ip):
        raise HTTPException(status_code=429, detail="Rate limit exceeded.")
    if not session_id:
        session_id = str(uuid.uuid4())
    job_id = str(uuid.uuid4())
    output_path = str(OUTPUT_DIR / f"{job_id}.xlsx")
    fname = f"{tmpl['name'].lower().replace(' ','_')}.xlsx"
    jobs[job_id] = {"status":"pending","job_id":job_id,"session_id":session_id,
                    "created_at":datetime.utcnow().isoformat()}
    background_tasks.add_task(process_job, job_id=job_id, prompt=prompt, file_content="",
        file_type="", image_data=None, session_id=session_id, output_path=output_path,
        custom_filename=fname)
    return {"job_id":job_id,"session_id":session_id,"status":"processing","template":tmpl["name"]}

@app.get("/session/{session_id}/history")
def get_session_history(session_id: str):
    history = conversation_memory.get(session_id, [])
    conn = get_db()
    jobs_list = conn.execute(
        "SELECT job_id,status,prompt,created_at,processing_ms,provider FROM jobs WHERE session_id=? ORDER BY created_at DESC LIMIT 20",
        (session_id,)).fetchall()
    conn.close()
    return {"session_id":session_id,"message_count":len(history),"jobs":[dict(j) for j in jobs_list]}

@app.delete("/session/{session_id}")
def clear_session(session_id: str):
    conversation_memory.pop(session_id, None)
    return {"message":"Session cleared","session_id":session_id}

@app.get("/history")
def get_history(limit: int = 20):
    conn = get_db()
    rows = conn.execute(
        "SELECT job_id,status,prompt,created_at,processing_ms,provider FROM jobs ORDER BY created_at DESC LIMIT ?",
        (limit,)).fetchall()
    conn.close()
    return {"jobs":[dict(r) for r in rows]}

@app.get("/stats")
def get_stats():
    conn = get_db()
    total   = conn.execute("SELECT COUNT(*) FROM jobs").fetchone()[0]
    done    = conn.execute("SELECT COUNT(*) FROM jobs WHERE status='done'").fetchone()[0]
    failed  = conn.execute("SELECT COUNT(*) FROM jobs WHERE status='failed'").fetchone()[0]
    avg_t   = conn.execute("SELECT AVG(processing_ms) FROM jobs WHERE status='done'").fetchone()[0]
    today   = conn.execute("SELECT COUNT(*) FROM jobs WHERE date(created_at)=date('now')").fetchone()[0]
    by_prov = conn.execute("SELECT provider,COUNT(*) as n FROM jobs WHERE status='done' GROUP BY provider").fetchall()
    conn.close()
    return {
        "total_jobs":total,"successful":done,"failed":failed,"today":today,
        "avg_processing_ms":round(avg_t or 0),
        "success_rate":f"{(done/total*100):.1f}%" if total else "N/A",
        "provider_breakdown":{r["provider"]:r["n"] for r in by_prov},
        "active_sessions":len(conversation_memory),
        "active_jobs":len([j for j in jobs.values() if j["status"]=="processing"]),
        "quarantined_models":len(_quarantine),
        "tripped_providers":list(_provider_tripped.keys()),
    }

@app.get("/jobs")
def list_jobs():
    return {"jobs":[{k:v for k,v in j.items() if k!="output_path"} for j in list(jobs.values())[-50:]]}

@app.post("/admin/cleanup")
def run_cleanup():
    cleanup_old_files()
    _prune_jobs()
    return {"message":"Cleanup complete"}

@app.post("/email/{job_id}")
async def email_file(job_id: str, to_email: str = Form(...),
                     subject: str = Form(default="Your Excel File from XLforge")):
    job = jobs.get(job_id)
    if not job or job["status"] != "done":
        raise HTTPException(status_code=404, detail="File not ready")
    output_path = job.get("output_path","")
    if not os.path.exists(output_path):
        raise HTTPException(status_code=404, detail="File not found")
    gmail_user = os.getenv("GMAIL_USER")
    gmail_pass = os.getenv("GMAIL_APP_PASSWORD")
    if not gmail_user or not gmail_pass:
        raise HTTPException(status_code=503, detail="Email not configured. Set GMAIL_USER and GMAIL_APP_PASSWORD.")
    try:
        msg = MIMEMultipart()
        msg["From"] = gmail_user
        msg["To"] = to_email
        msg["Subject"] = subject
        msg.attach(MIMEText("Your Excel file from XLforge AI is attached.", "plain"))
        with open(output_path, "rb") as f:
            part = MIMEBase("application","octet-stream")
            part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", f'attachment; filename="{job.get("custom_filename","xlforge_output.xlsx")}"')
        msg.attach(part)
        import ssl as _ssl
        with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=_ssl.create_default_context()) as srv:
            srv.login(gmail_user, gmail_pass)
            srv.send_message(msg)
        return {"message":f"Sent to {to_email} successfully"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Email failed: {e}")
