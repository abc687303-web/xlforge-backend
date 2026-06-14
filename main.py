from fastapi import FastAPI, UploadFile, File, Form, BackgroundTasks, HTTPException, Request, Depends
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse, StreamingResponse
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from typing import Optional, List
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, AreaChart, Reference
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, CellIsRule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.drawing.image import Image as XLImage
import os, uuid, httpx, json, re, io, zipfile, base64, csv, time, logging, asyncio
from datetime import datetime, timedelta
from pathlib import Path
import sqlite3
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders

# ======================================================
# SETUP & CONFIG
# ======================================================

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler("xlforge.log", mode="a")
    ]
)
logger = logging.getLogger("xlforge")

app = FastAPI(title="XLforge API", version="2.0.0")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Storage directories
STORAGE_DIR = Path("storage")
INPUT_DIR = STORAGE_DIR / "input"
OUTPUT_DIR = STORAGE_DIR / "output"
TEMP_DIR = STORAGE_DIR / "temp"
for d in [INPUT_DIR, OUTPUT_DIR, TEMP_DIR]:
    d.mkdir(parents=True, exist_ok=True)

# In-memory stores
conversation_memory: dict = {}   # session_id -> list of messages
rate_limit_store: dict = {}      # ip -> [timestamps]
jobs: dict = {}                  # job_id -> status dict

MAX_FILE_SIZE = 100 * 1024 * 1024  # 100MB
RATE_LIMIT = 15                     # requests per minute
FILE_EXPIRY_HOURS = 24


# ======================================================
# DATABASE
# ======================================================

def get_db():
    conn = sqlite3.connect("xlforge.db", check_same_thread=False)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db()
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS jobs (
            job_id TEXT PRIMARY KEY,
            status TEXT DEFAULT 'pending',
            prompt TEXT,
            session_id TEXT,
            input_file TEXT,
            output_file TEXT,
            error TEXT,
            created_at TEXT,
            completed_at TEXT,
            processing_ms INTEGER
        );
        CREATE TABLE IF NOT EXISTS templates (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT,
            category TEXT,
            prompt TEXT,
            icon TEXT,
            created_at TEXT
        );
        CREATE TABLE IF NOT EXISTS usage_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            job_id TEXT,
            session_id TEXT,
            ip TEXT,
            file_size INTEGER,
            ai_time_ms INTEGER,
            excel_time_ms INTEGER,
            success INTEGER,
            created_at TEXT
        );
    """)
    # Seed default templates
    existing = conn.execute("SELECT COUNT(*) FROM templates").fetchone()[0]
    if existing == 0:
        templates = [
            ("Invoice", "finance", "Create a professional invoice template with company details, line items (description, qty, unit price, total), subtotal, tax (18%), grand total, payment terms, bank details", "🧾"),
            ("Monthly Budget", "finance", "Create a personal monthly budget tracker with income sources, fixed expenses, variable expenses, savings goal, actual vs budget comparison, and summary chart", "💰"),
            ("Inventory Tracker", "business", "Create an inventory management sheet with product ID, name, category, quantity in stock, reorder level, unit cost, total value, supplier, last restocked date, and stock status alerts", "📦"),
            ("Employee Attendance", "HR", "Create an employee attendance sheet for a month with employee ID, name, department, 31 days columns (P/A/L/H), total present, total absent, total leave, attendance percentage", "👥"),
            ("Sales Report", "sales", "Create a monthly sales report with salesperson, region, product, units sold, unit price, revenue, target, achievement percentage, rank, and bar chart", "📈"),
            ("Project Timeline", "management", "Create a project timeline/gantt-style sheet with task name, owner, start date, end date, duration days, status (Not Started/In Progress/Complete), priority, dependencies, completion %", "📅"),
            ("Student Gradebook", "education", "Create a student gradebook with student name, roll number, 6 subjects scores, total, percentage, grade (A/B/C/D/F), rank, pass/fail status and class average row", "🎓"),
            ("Expense Report", "finance", "Create a business expense report with date, category, description, amount, receipt number, payment method, reimbursable yes/no, approval status, and monthly totals", "💳"),
            ("KPI Dashboard", "management", "Create a KPI dashboard with metrics (Revenue, Customers, Conversion Rate, Avg Order Value, Churn Rate, NPS), current value, previous period, target, variance, trend (↑↓), and status (On Track/At Risk/Behind)", "📊"),
            ("Quotation", "sales", "Create a business quotation template with client details, quote number, validity date, itemized list (item, specs, qty, unit price, discount, net price), terms & conditions, total, and signature section", "📋"),
            ("BOQ (Bill of Quantities)", "construction", "Create a BOQ sheet with item number, description of work, unit, quantity, rate, amount, GST %, GST amount, total amount, section subtotals and grand total", "🏗️"),
            ("Payroll Sheet", "HR", "Create a monthly payroll sheet with employee ID, name, designation, basic salary, HRA, DA, other allowances, gross salary, PF deduction, ESI, tax, total deductions, net salary", "💵"),
        ]
        conn.executemany(
            "INSERT INTO templates (name, category, prompt, icon, created_at) VALUES (?,?,?,?,?)",
            [(t[0], t[1], t[2], t[3], datetime.utcnow().isoformat()) for t in templates]
        )
    conn.commit()
    conn.close()

init_db()


# ======================================================
# RATE LIMITER
# ======================================================

def check_rate_limit(ip: str) -> bool:
    now = time.time()
    window = 60
    if ip not in rate_limit_store:
        rate_limit_store[ip] = []
    rate_limit_store[ip] = [t for t in rate_limit_store[ip] if now - t < window]
    if len(rate_limit_store[ip]) >= RATE_LIMIT:
        return False
    rate_limit_store[ip].append(now)
    return True


# ======================================================
# FILE READER
# ======================================================

async def read_any_file(file: UploadFile) -> tuple:
    raw = await file.read()

    if len(raw) > MAX_FILE_SIZE:
        raise HTTPException(status_code=413, detail=f"File too large. Max size is {MAX_FILE_SIZE // 1024 // 1024}MB per file")

    filename = (file.filename or "").lower()

    if filename.endswith(('.xlsx', '.xls')):
        try:
            wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
            lines = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                lines.append(f"[Sheet: {sheet_name}]")
                for row in ws.iter_rows(values_only=True):
                    if any(c is not None for c in row):
                        lines.append(" | ".join(str(c) if c is not None else "" for c in row))
            return "\n".join(lines), "excel", None, raw
        except Exception as e:
            return f"Excel read error: {e}", "excel", None, raw

    if filename.endswith('.csv'):
        try:
            return raw.decode("utf-8", errors="ignore"), "csv", None, raw
        except:
            return "", "csv", None, raw

    if filename.endswith('.docx'):
        try:
            z = zipfile.ZipFile(io.BytesIO(raw))
            xml = z.read("word/document.xml").decode("utf-8")
            text = re.sub(r'<[^>]+>', ' ', xml)
            text = re.sub(r'\s+', ' ', text).strip()
            return text[:6000], "word", None, raw
        except:
            return "", "word", None, raw

    if filename.endswith('.pdf'):
        try:
            text = raw.decode("latin-1", errors="ignore")
            strings = re.findall(r'[A-Za-z0-9 \+\-\=\.\,\:\;\!\?\%\$\#\@\/\(\)]{4,}', text)
            extracted = " ".join(strings)[:6000]
            return extracted, "pdf", None, raw
        except:
            return "", "pdf", None, raw

    if filename.endswith(('.jpg', '.jpeg', '.png', '.gif', '.webp', '.bmp')):
        b64 = base64.b64encode(raw).decode()
        ext = filename.split('.')[-1]
        mime = f"image/{'jpeg' if ext in ['jpg','jpeg'] else ext}"
        return "", "image", {"b64": b64, "mime": mime}, raw

    try:
        return raw.decode("utf-8", errors="ignore")[:6000], "text", None, raw
    except:
        return "", "unknown", None, raw


# ======================================================
# SYSTEM PROMPT
# ======================================================

SYSTEM_PROMPT = """You are XLforge, an AI Excel expert. Return ONLY a valid JSON object — no markdown, no explanation, nothing else.

JSON FORMAT (all keys required):
{"sheets":[{"name":"Sheet Name","headers":["Col A","Col B","Col C"],"rows":[["text",100,50.5]],"formulas":[{"cell":"D2","formula":"=IFERROR(SUM(B2:C2),0)","label":"Total"}],"summary_rows":[{"label":"TOTAL","col":1,"formula":"=IFERROR(SUM(B2:B100),0)"}],"conditional_formatting":[{"range":"B2:B20","type":"colorscale"}],"chart":{"type":"bar","title":"Chart Title","data_cols":[1],"category_col":0}}],"metadata":{"title":"Report Title","description":"What this does","author":"XLforge AI"}}

ABSOLUTE RULES:
1. NUMBERS: prices/scores/salaries MUST be int or float — WRONG:["John","50000"] CORRECT:["John",50000]
2. MATH: if file has "1021+707" in Problem col, Answer col gets integer 1728 (not string). Problem text stays unchanged.
3. SUMMARY ROWS: "col" is 0-indexed. headers=["A","B","C"] -> col 2 = column C. Always IFERROR.
4. CHARTS: type=bar/line/pie/area; data_cols and category_col are 0-indexed integers.
   - Add chart ONLY if user explicitly requests one.
   - "chart in 2nd sheet" -> add "sheet":1 inside chart object.
5. FORMULAS: always wrap in IFERROR. SUM range must start at row 2 (never row 1).
6. ROWS: copy every row from uploaded file — no skipping, no reducing.
7. PLACEHOLDERS: never use val1/item1/data1. Always real meaningful data, minimum 15 rows.
8. ANY LANGUAGE: understand and process prompts in any language.
9. SELF-CORRECT: if you are unsure, make a best guess and produce useful output."""


# ======================================================
# VALIDATION LAYER
# ======================================================

def validate_ai_response(data: dict) -> tuple[bool, str]:
    if not isinstance(data, dict):
        return False, "Response is not a dict"
    if "sheets" not in data:
        return False, "Missing 'sheets' key"
    if not data["sheets"]:
        return False, "Empty sheets list"
    for i, sheet in enumerate(data["sheets"]):
        if not sheet.get("headers"):
            return False, f"Sheet {i}: missing headers"
        if not sheet.get("rows"):
            return False, f"Sheet {i}: missing rows"
        if not isinstance(sheet["headers"], list):
            return False, f"Sheet {i}: headers must be a list"
        if not isinstance(sheet["rows"], list):
            return False, f"Sheet {i}: rows must be a list"
        header_len = len(sheet["headers"])
        for j, row in enumerate(sheet["rows"]):
            if not isinstance(row, list):
                return False, f"Sheet {i}, Row {j}: must be a list"
            # Pad short rows
            while len(row) < header_len:
                row.append("")
    return True, "OK"


def coerce_numeric(val):
    """Convert plain numeric strings to int/float.
    CRITICAL: Do NOT eval math expressions like '1021 + 707'.
    Those are question/problem text - they must stay as strings in Excel.
    Only convert pure number strings: '125' -> 125, '3.14' -> 3.14, '-50' -> -50."""
    if isinstance(val, (int, float)):
        return val
    if isinstance(val, str):
        v = val.strip()
        if not v:
            return val
        # Never touch Excel formulas
        if v.startswith('='):
            return val
        # Only convert if it is a plain number: digits, optional leading minus, optional decimal.
        # Rejects anything with spaces or operators (+, *, /, etc.) - those are text data.
        if re.match(r'^-?[\d]+(\.[\d]+)?$', v):
            try:
                if '.' in v:
                    return float(v)
                return int(v)
            except (ValueError, TypeError):
                pass
        # Also accept numbers with thousand-commas like "1,234" or "1,234.56"
        if re.match(r'^-?[\d]{1,3}(,[\d]{3})*(\.[\d]+)?$', v):
            try:
                cleaned = v.replace(',', '')
                if '.' in cleaned:
                    return float(cleaned)
                return int(cleaned)
            except (ValueError, TypeError):
                pass
    return val

def sanitize_ai_response(data: dict) -> dict:
    for sheet in data.get("sheets", []):
        headers = sheet.get("headers", [])
        header_len = len(headers)

        # Identify problem/question columns - these must stay as strings (e.g. "93 + 19")
        problem_col_indices = set()
        for i, h in enumerate(headers):
            hl = str(h).lower()
            if any(k in hl for k in ["problem", "question", "expression", "equation", "task"]):
                problem_col_indices.add(i)

        cleaned_rows = []
        for row in sheet.get("rows", []):
            if isinstance(row, list):
                row = row[:header_len]
                while len(row) < header_len:
                    row.append("")
                # Coerce numeric strings to real numbers, but protect problem/question columns
                new_row = []
                for ci, v in enumerate(row):
                    if ci in problem_col_indices:
                        # Keep as string - never coerce math expressions like "93 + 19"
                        new_row.append(str(v) if v is not None else "")
                    else:
                        new_row.append(coerce_numeric(v))
                cleaned_rows.append(new_row)
        sheet["rows"] = cleaned_rows
        if not sheet.get("formulas"):
            sheet["formulas"] = []
        if not sheet.get("conditional_formatting"):
            sheet["conditional_formatting"] = []
        if not sheet.get("summary_rows"):
            sheet["summary_rows"] = []
    return data


# ======================================================
# MODEL AUTO-DISCOVERY
# ======================================================

# Module-level cache: refreshed every hour so the app always uses active models
_model_cache: dict = {"text": None, "image": None, "expires": None}

# Models that handle structured text output well — tried in priority order
_TEXT_MODEL_PRIORITY = [
    "llama-3.3-70b-versatile",
    "qwen/qwen3-32b",
    "llama-3.1-8b-instant",
    "groq/compound-mini",
    "groq/compound",
    "openai/gpt-oss-20b",
    "openai/gpt-oss-120b",
    "allam-2-7b",
]

# Patterns in model IDs that indicate non-text models to skip
_NON_TEXT_PATTERNS = [
    "whisper", "tts", "speech", "guard", "orpheus",
    "canopylabs/", "prompt-guard",
]

# Vision models (multimodal) — kept separate since image uploads need them
_VISION_MODEL_PRIORITY = [
    "meta-llama/llama-4-scout-17b-16e-instruct",
    "llama-3.3-70b-versatile",
    "llama-3.1-8b-instant",
]


async def _fetch_groq_models(groq_api_key: str) -> list[str]:
    """Return list of all active model IDs from Groq /v1/models."""
    async with httpx.AsyncClient(timeout=10) as client:
        r = await client.get(
            "https://api.groq.com/openai/v1/models",
            headers={"Authorization": f"Bearer {groq_api_key}"},
        )
        r.raise_for_status()
        return [m["id"] for m in r.json().get("data", [])]


async def get_active_text_models(groq_api_key: str) -> list[str]:
    """
    Return up to 4 text-capable Groq models that are currently active,
    ordered by quality. Result is cached for 1 hour.
    Falls back to a known-good list if the API call fails.
    """
    global _model_cache
    now = datetime.utcnow()

    if _model_cache["text"] and _model_cache["expires"] and _model_cache["expires"] > now:
        return _model_cache["text"]

    try:
        all_ids = await _fetch_groq_models(groq_api_key)

        # Filter to text-capable models only
        text_ids = set(
            m for m in all_ids
            if not any(p in m.lower() for p in _NON_TEXT_PATTERNS)
        )

        # Build prioritised list: priority models first, then whatever else is active
        ordered = [m for m in _TEXT_MODEL_PRIORITY if m in text_ids]
        for m in sorted(text_ids):
            if m not in ordered:
                ordered.append(m)

        result = ordered[:4] if ordered else ["llama-3.3-70b-versatile", "llama-3.1-8b-instant"]
        _model_cache["text"] = result
        _model_cache["image"] = [m for m in _VISION_MODEL_PRIORITY if m in set(all_ids)][:3] or ["llama-3.3-70b-versatile"]
        _model_cache["expires"] = now + timedelta(hours=1)
        logger.info(f"Auto-discovered Groq text models: {result}")
        return result

    except Exception as e:
        logger.warning(f"Model discovery failed ({e}), using hardcoded fallback list")
        fallback = ["llama-3.3-70b-versatile", "llama-3.1-8b-instant"]
        _model_cache["text"] = fallback
        _model_cache["expires"] = now + timedelta(minutes=5)  # retry sooner on failure
        return fallback


async def get_active_image_models(groq_api_key: str) -> list[str]:
    """Return vision-capable models; triggers text model discovery if cache is cold."""
    await get_active_text_models(groq_api_key)  # populates _model_cache["image"]
    return _model_cache.get("image") or ["llama-3.3-70b-versatile"]


# ======================================================
# JSON SELF-REPAIR
# ======================================================

def repair_json(text: str) -> str:
    """
    Self-heal common AI JSON truncation issues:
    - Unterminated strings: close the string, then close open structures
    - Missing closing braces/brackets: use a stack to close in correct order
    - Trailing commas before ] or }
    """
    # Step 1: Detect and close any unterminated string
    in_string = False
    escape_next = False
    for ch in text:
        if escape_next:
            escape_next = False
            continue
        if ch == '\\' and in_string:
            escape_next = True
            continue
        if ch == '"':
            in_string = not in_string
    if in_string:
        text = text + '"'   # close the open string

    # Step 2: Remove trailing commas before ] or } and strip trailing whitespace
    text = re.sub(r',\s*([}\]])', r'\1', text)
    text = text.rstrip(' \t\n,')

    # Step 3: Walk the text with a stack to figure out what needs closing.
    # This correctly interleaves ] and } in the right order (unlike simple counting).
    stack = []
    in_str = False
    esc = False
    for ch in text:
        if esc:
            esc = False
            continue
        if ch == '\\' and in_str:
            esc = True
            continue
        if ch == '"':
            in_str = not in_str
            continue
        if in_str:
            continue
        if ch == '{':
            stack.append('}')
        elif ch == '[':
            stack.append(']')
        elif ch in ('}', ']') and stack:
            stack.pop()

    # Close all open structures in reverse order
    for closer in reversed(stack):
        text += closer

    # Step 4: Final trailing-comma cleanup after adding closers
    text = re.sub(r',\s*([}\]])', r'\1', text)
    return text


# ======================================================
# AI CALL (GROQ)
# ======================================================

async def call_groq(
    prompt: str,
    file_content: str = "",
    file_type: str = "",
    image_data: dict = None,
    session_id: str = None,
) -> dict:

    messages = [{"role": "system", "content": SYSTEM_PROMPT}]

    # Inject conversation history
    if session_id and session_id in conversation_memory:
        history = conversation_memory[session_id][-6:]  # last 3 exchanges
        messages.extend(history)

    # Build user message
    if image_data:
        messages.append({
            "role": "user",
            "content": [
                {
                    "type": "image_url",
                    "image_url": {"url": f"data:{image_data['mime']};base64,{image_data['b64']}"}
                },
                {
                    "type": "text",
                    "text": f"""Analyze this image carefully.
{"Task: " + prompt if prompt.strip() else "Understand this image and create the best Excel spreadsheet from it."}
Read all text, numbers, tables visible in the image.
Solve any math problems, complete any missing data.
Return only JSON."""
                }
            ]
        })
    elif file_content and not prompt.strip():
        messages.append({
            "role": "user",
            "content": f"""I uploaded a {file_type} file. Read every row and create a professional Excel output.

FILE CONTENT:
{file_content[:4000]}

- Solve math problems (put integer answers in Answer column, keep problem text unchanged)
- Add Grade/Total/Rank/Pass-Fail for student marks
- Copy every row without skipping
- Return only JSON"""
        })
    elif file_content and prompt.strip():
        messages.append({
            "role": "user",
            "content": f"""Task: {prompt}

FILE TYPE: {file_type}
FILE CONTENT:
{file_content[:4000]}

- Use every row from the file
- Only do exactly what the user asked (no extra sheets/charts unless requested)
- Return only JSON"""
        })
    else:
        messages.append({
            "role": "user",
            "content": f"""Create a professional Excel spreadsheet: "{prompt}"
- Real, meaningful data - minimum 15 rows
- Multiple sheets if it makes sense
- Professional formulas throughout
- Add chart ONLY if the prompt specifically asks for one or data is clearly comparative
- Add conditional formatting
- Summary row at bottom
- No placeholder values whatsoever
- STRICTLY follow only what the prompt asks for, nothing extra
- Return only JSON"""
        })

    last_error = None
    groq_api_key = os.getenv('GROQ_API_KEY')
    if not groq_api_key:
        raise ValueError("GROQ_API_KEY environment variable is not set.")

    # Auto-discover active models from Groq API (cached 1 hour, falls back on error)
    TEXT_MODELS = await get_active_text_models(groq_api_key)
    IMAGE_MODELS = await get_active_image_models(groq_api_key)
    # Pad to exactly 4 attempts by repeating the last available model
    while len(TEXT_MODELS) < 4:
        TEXT_MODELS = TEXT_MODELS + [TEXT_MODELS[-1]]

    # Content limits shrink each retry to stay within model context windows:
    CONTENT_LIMITS = [4000, 2000, 1000, 500]
    # max_tokens per attempt (input ~300 tokens + max_tokens should be < 6000):
    MAX_TOKENS_PER_ATTEMPT = [5500, 4500, 3500, 2500]
    model = "unknown"  # safety init — overwritten at start of each attempt
    for attempt in range(4):
        try:
            temperature = [0.1, 0.2, 0.35, 0.5][attempt]
            content_limit = CONTENT_LIMITS[attempt]

            if image_data:
                model = IMAGE_MODELS[min(attempt, len(IMAGE_MODELS)-1)]
                # If falling back to a text-only model, strip the image payload
                is_vision_model = "llama-4" in model
                if not is_vision_model and isinstance(messages[-1]["content"], list):
                    # Extract just the text part for text-only fallback
                    text_parts = [p["text"] for p in messages[-1]["content"] if p.get("type") == "text"]
                    messages[-1] = {"role": "user", "content": " ".join(text_parts)}
            else:
                model = TEXT_MODELS[min(attempt, len(TEXT_MODELS)-1)]

            # Rebuild user message with content truncated to this attempt's limit
            if not image_data and file_content and attempt > 0:
                truncated = file_content[:content_limit]
                if file_content and not prompt.strip():
                    messages[-1] = {
                        "role": "user",
                        "content": f"""I uploaded a {file_type} file. Read it and create a professional Excel output.

FILE CONTENT:
{truncated}

CRITICAL: Read every row, solve math if present, add totals. Return only JSON."""
                    }
                elif file_content and prompt.strip():
                    messages[-1] = {
                        "role": "user",
                        "content": f"""Task: {prompt}

FILE TYPE: {file_type}
FILE CONTENT:
{truncated}

Use every row from the file. Return only JSON."""
                    }

            async with httpx.AsyncClient(timeout=60) as client:
                response = await client.post(
                    "https://api.groq.com/openai/v1/chat/completions",
                    headers={
                        "Authorization": f"Bearer {groq_api_key}",
                        "Content-Type": "application/json"
                    },
                    json={
                        "model": model,
                        "max_tokens": MAX_TOKENS_PER_ATTEMPT[attempt],
                        "temperature": temperature,
                        "messages": messages
                    }
                )

            if response.status_code == 429:
                logger.warning(f"Rate limit on {model}, rotating to next model...")
                last_error = f"Rate limit on {model}"
                await asyncio.sleep(3 + attempt * 2)  # 3s, 5s, 7s, 9s progressive backoff
                continue
            if response.status_code == 413:
                logger.warning(f"Payload too large on {model} (attempt {attempt+1}), retrying with smaller content...")
                last_error = f"Groq HTTP 413: request too large on {model}"
                await asyncio.sleep(1)
                continue
            if response.status_code not in (200, 429):
                err_body = response.text[:300]
                # Permanent errors (400 bad model, 401 auth, 403 forbidden): log and skip model
                if response.status_code in (400, 401, 403, 404):
                    logger.error(f"Permanent error on {model} [{response.status_code}]: {err_body}")
                    last_error = f"Groq HTTP {response.status_code} on {model}: {err_body}"
                    continue  # try next model
                raise ValueError(f"Groq HTTP {response.status_code}: {err_body}")

            result = response.json()
            if "error" in result:
                raise ValueError(f"Groq error: {result['error']}")

            text = result["choices"][0]["message"]["content"].strip()
            text = re.sub(r"^```[a-zA-Z]*\n?", "", text)
            text = re.sub(r"\n?```$", "", text)
            text = text.strip()

            # Extract outermost JSON object — string-aware so "{}" inside string values
            # don't confuse the depth counter
            brace_start = text.find('{')
            if brace_start != -1:
                depth = 0
                brace_end = -1
                in_str = False
                esc = False
                for i, ch in enumerate(text[brace_start:], brace_start):
                    if esc:
                        esc = False
                        continue
                    if ch == '\\' and in_str:
                        esc = True
                        continue
                    if ch == '"':
                        in_str = not in_str
                        continue
                    if in_str:
                        continue
                    if ch == '{':
                        depth += 1
                    elif ch == '}':
                        depth -= 1
                        if depth == 0:
                            brace_end = i + 1
                            break
                text = text[brace_start:brace_end] if brace_end != -1 else text[brace_start:]

            # Always apply self-repair (fixes unterminated strings, trailing commas, open brackets)
            text = repair_json(text)

            try:
                data = json.loads(text)
            except json.JSONDecodeError:
                # Last-resort: strip everything after the last complete row array
                last_bracket = text.rfind(']]')
                if last_bracket != -1:
                    text = text[:last_bracket + 2]
                    text = repair_json(text)
                data = json.loads(text)

            valid, reason = validate_ai_response(data)
            if not valid:
                raise ValueError(f"Validation failed: {reason}")

            data = sanitize_ai_response(data)

            # Save to conversation memory
            if session_id:
                user_msg = {"role": "user", "content": prompt or f"[{file_type} file uploaded]"}
                assistant_msg = {"role": "assistant", "content": f"[Generated Excel with {len(data['sheets'])} sheet(s)]"}
                if session_id not in conversation_memory:
                    conversation_memory[session_id] = []
                conversation_memory[session_id].extend([user_msg, assistant_msg])
                conversation_memory[session_id] = conversation_memory[session_id][-20:]  # keep last 10 exchanges

            return data

        except json.JSONDecodeError as e:
            last_error = f"JSON parse error: {e}"
            logger.warning(f"Attempt {attempt+1} JSON error: {e}")
            continue
        except ValueError as e:
            last_error = str(e)
            logger.warning(f"Attempt {attempt+1} value error: {e}")
            continue
        except (httpx.TimeoutException, httpx.ReadTimeout, httpx.ConnectTimeout):
            last_error = "AI service timed out"
            logger.warning(f"Attempt {attempt+1} timeout on {model}")
            await asyncio.sleep(2)
            continue
        except Exception as e:
            last_error = str(e)
            logger.error(f"Attempt {attempt+1} unexpected error: {e}")
            continue

    raise ValueError(f"AI generation failed after 4 attempts. Last error: {last_error}")


# ======================================================
# EXCEL BUILDER
# ======================================================

def build_excel(data: dict, output_path: str, password: str = None):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    BLUE = "2563EB"
    LIGHT_BLUE = "EFF6FF"
    WHITE = "FFFFFF"
    DARK = "1E293B"
    SUMMARY_BG = "1E3A5F"
    pending_charts = []  # collected in pass 1, rendered in pass 2

    def thin_border():
        s = Side(style="thin", color="CCCCCC")
        return Border(left=s, right=s, top=s, bottom=s)

    def thick_border():
        s = Side(style="medium", color="2563EB")
        return Border(left=s, right=s, top=s, bottom=s)

    def style_header(cell):
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.font = Font(bold=True, color=WHITE, size=11, name="Calibri")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border()

    def style_data(cell, row_idx, header=""):
        cell.border = thin_border()
        cell.alignment = Alignment(vertical="center")
        if row_idx % 2 == 0:
            cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        if isinstance(cell.value, (int, float)):
            cell.alignment = Alignment(horizontal="right", vertical="center")
            h = header.lower()
            if any(k in h for k in ["salary","revenue","price","cost","amount","budget",
                                      "sales","income","expense","pay","total","value","rate"]):
                cell.number_format = '#,##0.00'
            elif any(k in h for k in ["percent","%","rate","growth","margin","achievement"]):
                cell.number_format = '0.00'

    def style_summary(cell):
        cell.fill = PatternFill("solid", fgColor=SUMMARY_BG)
        cell.font = Font(bold=True, color=WHITE, size=11, name="Calibri")
        cell.border = thick_border()
        cell.alignment = Alignment(horizontal="right", vertical="center")

    for sheet_def in data["sheets"]:
        name = str(sheet_def.get("name", "Sheet"))[:31]
        ws = wb.create_sheet(title=name)

        headers  = sheet_def.get("headers", [])
        rows     = sheet_def.get("rows", [])
        formulas = sheet_def.get("formulas", [])
        cf_rules = sheet_def.get("conditional_formatting", [])
        chart_def = sheet_def.get("chart", None)
        summary_rows = sheet_def.get("summary_rows", [])

        # -- Tab color
        ws.sheet_properties.tabColor = BLUE

        # -- Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=str(header))
            style_header(cell)
            ws.column_dimensions[get_column_letter(col)].width = max(len(str(header)) + 6, 18)
        ws.row_dimensions[1].height = 32

        # -- Write data rows
        for row_idx, row in enumerate(rows, 2):
            for col_idx, val in enumerate(row, 1):
                header = headers[col_idx-1] if col_idx <= len(headers) else ""
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                style_data(cell, row_idx, header)
            ws.row_dimensions[row_idx].height = 20

        # -- Auto-fit column widths based on actual data content
        for col_idx, header in enumerate(headers, 1):
            max_len = len(str(header))
            for row in rows:
                if col_idx - 1 < len(row):
                    cell_val = row[col_idx - 1]
                    max_len = max(max_len, len(str(cell_val)) if cell_val is not None else 0)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 4, 14), 60)

        # -- Summary rows at bottom - AUTO-GENERATED, never trust AI col index
        next_row = len(rows) + 2
        data_start_row = 2
        data_end_row = len(rows) + 1

        # Detect which columns are numeric by checking actual data
        numeric_cols = set()
        for row in rows:
            for ci, val in enumerate(row):
                if isinstance(val, (int, float)):
                    numeric_cols.add(ci)  # 0-indexed

        # Write TOTAL row if there are numeric columns
        if numeric_cols:
            for col_0idx in range(len(headers)):
                cell = ws.cell(row=next_row, column=col_0idx + 1)
                if col_0idx == 0:
                    cell.value = "TOTAL"
                elif col_0idx in numeric_cols:
                    col_letter = get_column_letter(col_0idx + 1)
                    cell.value = f'=IFERROR(SUM({col_letter}{data_start_row}:{col_letter}{data_end_row}),0)'
                else:
                    cell.value = None
                style_summary(cell)
            next_row += 1

            # Write AVERAGE row - IFERROR prevents #DIV/0! on empty/formula-only ranges
            for col_0idx in range(len(headers)):
                cell = ws.cell(row=next_row, column=col_0idx + 1)
                if col_0idx == 0:
                    cell.value = "AVERAGE"
                elif col_0idx in numeric_cols:
                    col_letter = get_column_letter(col_0idx + 1)
                    cell.value = f'=IFERROR(AVERAGE({col_letter}{data_start_row}:{col_letter}{data_end_row}),0)'
                else:
                    cell.value = None
                style_summary(cell)
            next_row += 1

        # -- Inline formulas
        for f in formulas:
            addr = f.get("cell", "")
            formula = f.get("formula", "")
            label = f.get("label", "")
            if addr and formula:
                try:
                    ws[addr] = formula
                    ws[addr].font = Font(bold=True, color=DARK)
                    ws[addr].border = thick_border()
                    m = re.match(r"([A-Z]+)(\d+)", addr)
                    if m and label:
                        ci = column_index_from_string(m.group(1))
                        ri = int(m.group(2))
                        if ci > 1:
                            # Only write label if cell to the left has no data value
                            # (never overwrite real data like Volume, Price, etc.)
                            left_cell = ws.cell(row=ri, column=ci-1)
                            if left_cell.value is None or left_cell.value == "":
                                left_cell.value = label
                                left_cell.font = Font(bold=True)
                                left_cell.border = thin_border()
                except:
                    pass

        # -- Freeze panes & auto-filter
        ws.freeze_panes = "A2"
        if headers:
            last_data_row = len(rows) + 1
            ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{last_data_row}"

        # -- Print settings
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.sheet_view.showGridLines = True

        # -- Conditional formatting
        for cf in cf_rules:
            r = cf.get("range", "")
            t = cf.get("type", "")
            if not r:
                continue
            try:
                if t == "colorscale":
                    ws.conditional_formatting.add(r, ColorScaleRule(
                        start_type="min",  start_color="F87171",
                        mid_type="percentile", mid_value=50, mid_color="FCD34D",
                        end_type="max",    end_color="4ADE80"
                    ))
                elif t == "databar":
                    ws.conditional_formatting.add(r, DataBarRule(
                        start_type="min", start_value=0,
                        end_type="max",   end_value=100,
                        color="2563EB"
                    ))
                elif t == "highlight_high":
                    threshold = cf.get("threshold", 0)
                    ws.conditional_formatting.add(r, CellIsRule(
                        operator="greaterThan",
                        formula=[str(threshold)],
                        stopIfTrue=True,
                        fill=PatternFill(fill_type="solid", fgColor="4ADE80"),
                        font=Font(color="14532D", bold=True)
                    ))
                elif t == "highlight_low":
                    threshold = cf.get("threshold", 0)
                    ws.conditional_formatting.add(r, CellIsRule(
                        operator="lessThan",
                        formula=[str(threshold)],
                        stopIfTrue=True,
                        fill=PatternFill(fill_type="solid", fgColor="FCA5A5"),
                        font=Font(color="7F1D1D", bold=True)
                    ))
            except Exception as e:
                logger.warning(f"CF error: {e}")

        # -- Collect chart definition for second-pass rendering (after all sheets exist)
        if chart_def:
            pending_charts.append({
                "chart_def": chart_def,
                "data_ws": ws,
                "nrows": len(rows),
                "next_row": next_row,
            })

    # ======================================================
    # SECOND PASS: Add all charts now that all sheets exist
    # ======================================================
    for pc in pending_charts:
        try:
            chart_def    = pc["chart_def"]
            data_ws      = pc["data_ws"]
            nrows        = pc["nrows"]
            next_row_val = pc["next_row"]

            ctype  = str(chart_def.get("type", "bar")).lower()
            dcols  = chart_def.get("data_cols", [1])
            catcol = int(chart_def.get("category_col", 0))
            ctitle = str(chart_def.get("title", "Chart"))
            target_sheet_idx = chart_def.get("sheet", None)

            # Validate dcols
            if not isinstance(dcols, list) or not dcols:
                dcols = [1]
            dcols = [int(d) for d in dcols]

            # Build chart object with fallback
            if ctype == "pie":
                chart = PieChart()
            elif ctype == "line":
                chart = LineChart()
            elif ctype == "area":
                chart = AreaChart()
            else:
                chart = BarChart()
                chart.type = "col"

            chart.title  = ctitle
            chart.style  = 10
            chart.height = 15
            chart.width  = 28

            # Only reference actual data rows (exclude summary rows)
            data_max_row = nrows + 1  # row 1 = header, row 2..nrows+1 = data

            if ctype == "pie":
                dc = dcols[0]
                if dc + 1 > len(data_ws[1]):
                    dc = 1
                data_ref = Reference(data_ws, min_col=dc + 1, min_row=1, max_row=data_max_row)
                chart.add_data(data_ref, titles_from_data=True)
            else:
                for dc in dcols:
                    if dc + 1 > 50:
                        continue
                    data_ref = Reference(data_ws, min_col=dc + 1, min_row=1, max_row=data_max_row)
                    chart.add_data(data_ref, titles_from_data=True)

            cats = Reference(data_ws, min_col=catcol + 1, min_row=2, max_row=data_max_row)
            chart.set_categories(cats)

            # Place on target sheet (all sheets now exist)
            all_sheets = wb.worksheets
            if target_sheet_idx is not None and 0 <= int(target_sheet_idx) < len(all_sheets):
                chart_ws = all_sheets[int(target_sheet_idx)]
                anchor_row = 3
            else:
                chart_ws = data_ws
                anchor_row = max(nrows + 4, next_row_val + 2)

            chart_ws.add_chart(chart, f"A{anchor_row}")
            logger.info(f"Chart '{ctitle}' ({ctype}) placed on sheet '{chart_ws.title}' at row {anchor_row}")
        except Exception as e:
            logger.warning(f"Chart rendering error (skipped): {e}")

    wb.save(output_path)
    return output_path


# ======================================================
# CSV CONVERTER
# ======================================================

def excel_to_csv(excel_path: str) -> str:
    """Export first sheet of Excel to CSV. Returns the CSV file path."""
    csv_path = excel_path.replace(".xlsx", ".csv")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    # Use first sheet (wb.active may point to wrong sheet if active was changed)
    ws = wb.worksheets[0] if wb.worksheets else wb.active
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        for row in ws.iter_rows(values_only=True):
            writer.writerow([c if c is not None else "" for c in row])
    return csv_path


# ======================================================
# BACKGROUND JOB PROCESSOR
# ======================================================

async def process_job(
    job_id: str,
    prompt: str,
    file_content: str,
    file_type: str,
    image_data: dict,
    session_id: str,
    output_path: str,
    password: str = None,
    custom_filename: str = None,
):
    start = time.time()
    conn = get_db()
    try:
        jobs[job_id]["status"] = "processing"
        conn.execute("UPDATE jobs SET status='processing' WHERE job_id=?", (job_id,))
        conn.commit()

        ai_start = time.time()
        data = await call_groq(prompt, file_content, file_type, image_data, session_id)
        ai_ms = int((time.time() - ai_start) * 1000)

        excel_start = time.time()
        build_excel(data, output_path, password=password)
        excel_ms = int((time.time() - excel_start) * 1000)

        total_ms = int((time.time() - start) * 1000)

        jobs[job_id].update({
            "status": "done",
            "output_path": output_path,
            "custom_filename": custom_filename or "xlforge_output.xlsx",
            "sheet_count": len(data.get("sheets", [])),
            "metadata": data.get("metadata", {}),
            "processing_ms": total_ms
        })

        conn.execute("""
            UPDATE jobs SET status='done', output_file=?, completed_at=?, processing_ms=?
            WHERE job_id=?
        """, (output_path, datetime.utcnow().isoformat(), total_ms, job_id))
        conn.execute("""
            INSERT INTO usage_log (job_id, session_id, ai_time_ms, excel_time_ms, success, created_at)
            VALUES (?,?,?,?,1,?)
        """, (job_id, session_id, ai_ms, excel_ms, datetime.utcnow().isoformat()))
        conn.commit()

        logger.info(f"Job {job_id} done in {total_ms}ms (AI:{ai_ms}ms, Excel:{excel_ms}ms)")

    except Exception as e:
        jobs[job_id].update({"status": "failed", "error": str(e)})
        try:
            conn.execute("UPDATE jobs SET status='failed', error=? WHERE job_id=?", (str(e), job_id))
            conn.execute("""
                INSERT INTO usage_log (job_id, session_id, success, created_at)
                VALUES (?,?,0,?)
            """, (job_id, session_id, datetime.utcnow().isoformat()))
            conn.commit()
        except Exception as db_err:
            logger.error(f"DB write failed for job {job_id}: {db_err}")
        logger.error(f"Job {job_id} failed: {e}")
    finally:
        conn.close()


# ======================================================
# FILE CLEANUP (run periodically)
# ======================================================

def cleanup_old_files():
    cutoff = time.time() - (FILE_EXPIRY_HOURS * 3600)
    for folder in [INPUT_DIR, OUTPUT_DIR, TEMP_DIR]:
        for f in folder.glob("*"):
            try:
                if f.stat().st_mtime < cutoff:
                    f.unlink()
                    logger.info(f"Cleaned up: {f}")
            except Exception as e:
                logger.debug(f"Cleanup skip {f}: {e}")


def _prune_jobs_dict():
    """Remove completed/failed jobs older than 24 h from the in-memory dict to prevent unbounded growth."""
    cutoff = datetime.utcnow() - timedelta(hours=24)
    to_remove = []
    for jid, job in list(jobs.items()):
        if job.get("status") in ("done", "failed"):
            try:
                if datetime.fromisoformat(job.get("created_at", "")) < cutoff:
                    to_remove.append(jid)
            except (ValueError, TypeError):
                pass
    for jid in to_remove:
        jobs.pop(jid, None)
    if to_remove:
        logger.info(f"Pruned {len(to_remove)} old jobs from memory")


# ======================================================
# ENDPOINTS
# ======================================================

@app.get("/")
def root():
    return {
        "status": "XLforge v2.0 running!",
        "features": ["async jobs", "conversation memory", "templates", "CSV export",
                      "multiple files", "rate limiting", "history", "validation"]
    }

@app.get("/health")
def health():
    return {
        "status": "healthy",
        "active_jobs": len([j for j in jobs.values() if j["status"] == "processing"]),
        "sessions": len(conversation_memory),
        "timestamp": datetime.utcnow().isoformat()
    }


@app.get("/models")
async def list_models():
    """Return the currently active Groq models being used, and refresh the cache."""
    groq_api_key = os.getenv("GROQ_API_KEY")
    if not groq_api_key:
        raise HTTPException(status_code=503, detail="GROQ_API_KEY not configured")
    # Force refresh by clearing the cache expiry
    _model_cache["expires"] = None
    text_models = await get_active_text_models(groq_api_key)
    image_models = await get_active_image_models(groq_api_key)
    return {
        "text_models": text_models,
        "image_models": image_models,
        "cache_expires": _model_cache["expires"].isoformat() if _model_cache["expires"] else None,
        "note": "text_models[0] is attempt-1 (best quality); fallback order left→right"
    }


@app.on_event("startup")
async def _on_startup():
    """At startup: pre-warm the Groq model cache and launch the hourly cleanup loop."""
    # 1. Pre-warm model discovery so the first user request has no extra latency
    groq_api_key = os.getenv("GROQ_API_KEY")
    if groq_api_key:
        try:
            models = await get_active_text_models(groq_api_key)
            logger.info(f"Startup model pre-warm complete: {models}")
        except Exception as e:
            logger.warning(f"Startup model pre-warm failed: {e}")

    # 2. Schedule hourly file and memory cleanup so the server never leaks
    async def _cleanup_loop():
        while True:
            await asyncio.sleep(3600)
            try:
                cleanup_old_files()
                _prune_jobs_dict()
            except Exception as e:
                logger.warning(f"Scheduled cleanup error: {e}")

    asyncio.create_task(_cleanup_loop())


# -- MAIN GENERATE ENDPOINT (async job)
@app.post("/generate")
async def generate_excel(
    background_tasks: BackgroundTasks,
    request: Request,
    prompt: str = Form(default=""),
    session_id: str = Form(default=""),
    custom_filename: str = Form(default=""),
    password: str = Form(default=""),
    files: List[UploadFile] = File(default=[]),
    file: UploadFile = File(default=None),
):
    ip = request.client.host
    if not check_rate_limit(ip):
        raise HTTPException(status_code=429, detail="Rate limit exceeded. Max 15 requests/minute.")

    if not session_id:
        session_id = str(uuid.uuid4())

    job_id = str(uuid.uuid4())
    output_path = str(OUTPUT_DIR / f"{job_id}.xlsx")

    # Handle single or multiple files
    all_files = [uf for uf in (files or []) if uf and uf.filename]
    if file and file.filename:
        all_files.insert(0, file)

    file_content = ""
    file_type = ""
    image_data = None

    if all_files:
        # Merge content from all files
        contents = []
        for uf in all_files[:20]:  # max 20 files
            fc, ft, img, raw = await read_any_file(uf)
            if img:
                image_data = img  # use last image
            elif fc:
                contents.append(f"[File: {uf.filename}]\n{fc}")
            file_type = ft

            # Save input file
            input_path = INPUT_DIR / f"{job_id}_{uf.filename}"
            with open(str(input_path), "wb") as fout:
                fout.write(raw)

        file_content = "\n\n".join(contents)

    # Create job record
    jobs[job_id] = {
        "status": "pending",
        "job_id": job_id,
        "session_id": session_id,
        "created_at": datetime.utcnow().isoformat()
    }

    conn = get_db()
    conn.execute("""
        INSERT INTO jobs (job_id, status, prompt, session_id, created_at)
        VALUES (?,?,?,?,?)
    """, (job_id, "pending", prompt, session_id, datetime.utcnow().isoformat()))
    conn.commit()
    conn.close()

    fname = custom_filename.strip() or "xlforge_output"
    if not fname.endswith(".xlsx"):
        fname += ".xlsx"

    background_tasks.add_task(
        process_job,
        job_id=job_id,
        prompt=prompt,
        file_content=file_content,
        file_type=file_type,
        image_data=image_data,
        session_id=session_id,
        output_path=output_path,
        password=password or None,
        custom_filename=fname
    )

    return {
        "job_id": job_id,
        "session_id": session_id,
        "status": "processing",
        "message": "Your Excel is being generated. Poll /status/{job_id} to check progress."
    }


# -- STATUS CHECK
@app.get("/status/{job_id}")
def job_status(job_id: str):
    job = jobs.get(job_id)
    if not job:
        conn = get_db()
        row = conn.execute("SELECT * FROM jobs WHERE job_id=?", (job_id,)).fetchone()
        conn.close()
        if not row:
            raise HTTPException(status_code=404, detail="Job not found")
        return dict(row)
    return job


# -- DOWNLOAD EXCEL
@app.get("/download/{job_id}")
def download_excel(job_id: str):
    job = jobs.get(job_id)
    if not job or job["status"] != "done":
        raise HTTPException(status_code=404, detail="File not ready or job not found")
    output_path = job.get("output_path", "")
    if not os.path.exists(output_path):
        raise HTTPException(status_code=404, detail="File not found on disk")
    filename = job.get("custom_filename", "xlforge_output.xlsx")
    return FileResponse(
        output_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=filename,
        headers={"Access-Control-Allow-Origin": "*"}
    )


# -- DOWNLOAD AS CSV
@app.get("/download-csv/{job_id}")
def download_csv(job_id: str):
    job = jobs.get(job_id)
    if not job or job["status"] != "done":
        raise HTTPException(status_code=404, detail="File not ready")
    output_path = job.get("output_path", "")
    if not os.path.exists(output_path):
        raise HTTPException(status_code=404, detail="File not found")
    csv_path = excel_to_csv(output_path)
    filename = job.get("custom_filename", "xlforge_output.xlsx").replace(".xlsx", ".csv")
    return FileResponse(
        csv_path,
        media_type="text/csv",
        filename=filename,
        headers={"Access-Control-Allow-Origin": "*"}
    )


# -- SYNC GENERATE (for simple/quick requests, backwards compatible)
@app.post("/generate-sync")
async def generate_sync(
    request: Request,
    prompt: str = Form(default=""),
    session_id: str = Form(default=""),
    custom_filename: str = Form(default=""),
    file: UploadFile = File(None)
):
    ip = request.client.host
    if not check_rate_limit(ip):
        raise HTTPException(status_code=429, detail="Rate limit exceeded.")

    if not session_id:
        session_id = str(uuid.uuid4())

    job_id = str(uuid.uuid4())
    output_path = str(OUTPUT_DIR / f"{job_id}.xlsx")

    file_content, file_type, image_data = "", "", None
    if file and file.filename:
        file_content, file_type, image_data, _ = await read_any_file(file)

    try:
        data = await call_groq(prompt, file_content, file_type, image_data, session_id)
        build_excel(data, output_path)
    except Exception as e:
        try:
            Path(output_path).unlink(missing_ok=True)
        except Exception:
            pass
        raise HTTPException(status_code=500, detail=str(e))

    fname = custom_filename.strip() or "xlforge_output"
    if not fname.endswith(".xlsx"):
        fname += ".xlsx"

    return FileResponse(
        output_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=fname,
        headers={"Access-Control-Allow-Origin": "*", "X-Session-Id": session_id}
    )


# -- TEMPLATES
@app.get("/templates")
def get_templates(category: str = None):
    conn = get_db()
    if category:
        rows = conn.execute(
            "SELECT * FROM templates WHERE category=? ORDER BY name", (category,)
        ).fetchall()
    else:
        rows = conn.execute("SELECT * FROM templates ORDER BY category, name").fetchall()
    conn.close()
    return {"templates": [dict(r) for r in rows]}

@app.get("/templates/categories")
def get_categories():
    conn = get_db()
    rows = conn.execute("SELECT DISTINCT category FROM templates ORDER BY category").fetchall()
    conn.close()
    return {"categories": [r["category"] for r in rows]}

@app.post("/templates/use/{template_id}")
async def use_template(
    template_id: int,
    background_tasks: BackgroundTasks,
    request: Request,
    session_id: str = Form(default=""),
    custom_data: str = Form(default=""),
):
    conn = get_db()
    tmpl = conn.execute("SELECT * FROM templates WHERE id=?", (template_id,)).fetchone()
    conn.close()
    if not tmpl:
        raise HTTPException(status_code=404, detail="Template not found")

    prompt = tmpl["prompt"]
    if custom_data.strip():
        prompt += f"\n\nAdditional data/context: {custom_data}"

    ip = request.client.host
    if not check_rate_limit(ip):
        raise HTTPException(status_code=429, detail="Rate limit exceeded.")

    if not session_id:
        session_id = str(uuid.uuid4())

    job_id = str(uuid.uuid4())
    output_path = str(OUTPUT_DIR / f"{job_id}.xlsx")
    fname = f"{tmpl['name'].lower().replace(' ','_')}.xlsx"

    jobs[job_id] = {"status": "pending", "job_id": job_id, "session_id": session_id,
                    "created_at": datetime.utcnow().isoformat()}

    background_tasks.add_task(
        process_job,
        job_id=job_id, prompt=prompt, file_content="", file_type="",
        image_data=None, session_id=session_id, output_path=output_path,
        custom_filename=fname
    )

    return {"job_id": job_id, "session_id": session_id, "status": "processing",
            "template": tmpl["name"]}


# -- CONVERSATION / SESSION
@app.get("/session/{session_id}/history")
def get_session_history(session_id: str):
    history = conversation_memory.get(session_id, [])
    conn = get_db()
    jobs_list = conn.execute(
        "SELECT job_id, status, prompt, created_at, processing_ms FROM jobs WHERE session_id=? ORDER BY created_at DESC LIMIT 20",
        (session_id,)
    ).fetchall()
    conn.close()
    return {
        "session_id": session_id,
        "message_count": len(history),
        "jobs": [dict(j) for j in jobs_list]
    }

@app.delete("/session/{session_id}")
def clear_session(session_id: str):
    if session_id in conversation_memory:
        del conversation_memory[session_id]
    return {"message": "Session cleared", "session_id": session_id}


# -- JOB HISTORY
@app.get("/history")
def get_history(limit: int = 20):
    conn = get_db()
    rows = conn.execute(
        "SELECT job_id, status, prompt, created_at, processing_ms FROM jobs ORDER BY created_at DESC LIMIT ?",
        (limit,)
    ).fetchall()
    conn.close()
    return {"jobs": [dict(r) for r in rows]}


# -- EMAIL ENDPOINT
@app.post("/email/{job_id}")
async def email_file(job_id: str, to_email: str = Form(...), subject: str = Form(default="Your Excel File from XLforge")):
    job = jobs.get(job_id)
    if not job or job["status"] != "done":
        raise HTTPException(status_code=404, detail="File not ready")

    output_path = job.get("output_path", "")
    if not os.path.exists(output_path):
        raise HTTPException(status_code=404, detail="File not found")

    gmail_user = os.getenv("GMAIL_USER")
    gmail_pass = os.getenv("GMAIL_APP_PASSWORD")
    if not gmail_user or not gmail_pass:
        raise HTTPException(status_code=503, detail="Email not configured. Set GMAIL_USER and GMAIL_APP_PASSWORD env vars.")

    try:
        msg = MIMEMultipart()
        msg["From"] = gmail_user
        msg["To"] = to_email
        msg["Subject"] = subject
        msg.attach(MIMEText("Please find your Excel file generated by XLforge AI attached.", "plain"))

        with open(output_path, "rb") as f:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(f.read())
        encoders.encode_base64(part)
        filename = job.get("custom_filename", "xlforge_output.xlsx")
        part.add_header("Content-Disposition", f'attachment; filename="{filename}"')
        msg.attach(part)

        import ssl as _ssl
        with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=_ssl.create_default_context()) as server:
            server.login(gmail_user, gmail_pass)
            server.send_message(msg)

        return {"message": f"File sent to {to_email} successfully"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Email failed: {str(e)}")


# -- STATS / MONITORING
@app.get("/stats")
def get_stats():
    conn = get_db()
    total = conn.execute("SELECT COUNT(*) FROM jobs").fetchone()[0]
    done = conn.execute("SELECT COUNT(*) FROM jobs WHERE status='done'").fetchone()[0]
    failed = conn.execute("SELECT COUNT(*) FROM jobs WHERE status='failed'").fetchone()[0]
    avg_time = conn.execute("SELECT AVG(processing_ms) FROM jobs WHERE status='done'").fetchone()[0]
    today = conn.execute(
        "SELECT COUNT(*) FROM jobs WHERE date(created_at)=date('now')"
    ).fetchone()[0]
    conn.close()
    return {
        "total_jobs": total,
        "successful": done,
        "failed": failed,
        "today": today,
        "avg_processing_ms": round(avg_time or 0),
        "success_rate": f"{(done/total*100):.1f}%" if total > 0 else "N/A",
        "active_sessions": len(conversation_memory),
        "active_jobs": len([j for j in jobs.values() if j["status"] == "processing"])
    }


# -- CLEANUP ENDPOINT
@app.post("/admin/cleanup")
def run_cleanup():
    cleanup_old_files()
    return {"message": "Cleanup complete"}


# -- LIST ACTIVE JOBS
@app.get("/jobs")
def list_jobs():
    return {
        "jobs": [
            {k: v for k, v in j.items() if k != "output_path"}
            for j in list(jobs.values())[-50:]
        ]
    }


# -- /process ALIAS (frontend compatibility — same as /generate)
@app.post("/process")
async def process_alias(
    background_tasks: BackgroundTasks,
    request: Request,
    prompt: str = Form(default=""),
    session_id: str = Form(default=""),
    custom_filename: str = Form(default=""),
    password: str = Form(default=""),
    files: List[UploadFile] = File(default=[]),
    file: UploadFile = File(default=None),
):
    return await generate_excel(
        background_tasks=background_tasks,
        request=request,
        prompt=prompt,
        session_id=session_id,
        custom_filename=custom_filename,
        password=password,
        files=files,
        file=file,
    )
